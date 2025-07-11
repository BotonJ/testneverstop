# inject_modules/text_renderer.py

from jinja2 import Environment, FileSystemLoader, Undefined
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import pandas as pd
import logging

class StrictUndefined(Undefined):
    def __str__(self):
        raise ValueError(f"Template variable '{self._undefined_name}' is not defined.")

def render_text_template_from_mapping(mapping_path, summary_values, alias_dict=None):
    df = pd.read_excel(mapping_path, sheet_name="text_mapping")
    df_template = df[df["字段名"] == "文字模板"]
    if df_template.empty:
        logging.warning("No '文字模板' found in text_mapping sheet.")
        return ""
    cell_value = df_template.iloc[0]["模板"]
    if not isinstance(cell_value, str):
        return ""

    env = Environment(undefined=StrictUndefined)
    tmpl = env.from_string(cell_value)
    
    # 为了模板渲染更健壮，将所有None值替换为空字符串
    cleaned_summary = {k: (v if v is not None else "") for k, v in summary_values.items()}
    
    try:
        rendered_text = tmpl.render(**cleaned_summary)
        return rendered_text
    except ValueError as e:
        logging.error(f"【模板渲染失败: {e}】 with summary_values: {cleaned_summary}")
        return f"【模板渲染失败: {e}】"


def inject_text_to_excel(wb_or_path, sheet_name="汇总区块", cell="K1", text=""):
    """
    Injects rendered text into a specified cell in an Excel file.
    核心修复：可以接收一个文件路径(str)或一个已加载的工作簿对象(Workbook)。
    """
    wb = None
    try:
        # --- 核心修复 ---
        if isinstance(wb_or_path, Workbook):
            wb = wb_or_path
        elif isinstance(wb_or_path, str):
            wb = load_workbook(wb_or_path)
        else:
            raise TypeError(f"expected str (file path) or Workbook object, not {type(wb_or_path)}")
        # --- 修复结束 ---

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            logging.warning(f"在工作簿中未找到名为 '{sheet_name}' 的Sheet，无法注入文字。")
            return

        ws[cell] = text
        logging.info(f"文字已成功准备注入到 {sheet_name}!{cell}")

        # 如果传入的是路径，则保存。如果传入的是对象，则由调用者负责保存。
        if isinstance(wb_or_path, str):
            wb.save(wb_or_path)

    except Exception as e:
        logging.error(f"注入文字到 {sheet_name}!{cell} 时出错: {e}")

def inject_summary_values_debug(file_path, summary_values):
    # This function remains unchanged, assuming it's for debugging and can work on a saved file.
    try:
        wb = load_workbook(file_path)
        sheet_name = "元数据"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        
        ws.cell(1, 1, "字段名")
        ws.cell(1, 2, "对应值")
        for idx, (k, v) in enumerate(summary_values.items()):
            ws.cell(idx + 2, 1, k)
            ws.cell(idx + 2, 2, v)
        wb.save(file_path)
    except Exception as e:
        logging.error(f"注入调试信息时出错: {e}")