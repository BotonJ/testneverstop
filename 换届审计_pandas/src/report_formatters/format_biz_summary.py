# /src/report_formatters/format_biz_summary.py

import pandas as pd
import logging
import re

def create_and_inject_biz_summary(prebuilt_wb, wb_to_fill, mapping_configs, year_header_formatter={}):
    """
    【V3 - 标准化适配修复版】
    - 修复了因预制件格式变更导致的 IndexError。
    - 现在从标准化的三列预制件中，通过正确的列索引来提取数据。
    """
    logging.info("  -> 生成并注入 '收入汇总' 和 '支出汇总' Sheet...")
    try:
        # 1. 读取科目配置
        df_subjects = mapping_configs['业务活动表汇总注入配置']
        income_subjects = df_subjects[df_subjects['类型'] == '收入']['科目名称'].tolist()
        expense_subjects = df_subjects[df_subjects['类型'] == '支出']['科目名称'].tolist()
        
        # 2. 从预制件中收集数据
        all_data = []
        
        # --- [核心修改] ---
        # 定义清晰的列索引常量，以提高可读性和健壮性
        COL_SUBJECT = 0      # A列: 项目
        COL_FINAL_AMOUNT = 2 # C列: 期末金额 (这修复了BUG)
        
        for sheet_name in prebuilt_wb.sheetnames:
            if '业务活动表' in sheet_name:
                year_match = re.search(r'(\d{4})', sheet_name)
                if not year_match: continue
                year_int = int(year_match.group(0))
                
                ws = prebuilt_wb[sheet_name]
                # 从第2行开始遍历标准化的预制件数据
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 安全地检查元组长度，防止空行导致的问题
                    if len(row) < 3: continue
                        
                    subject = row[COL_SUBJECT]
                    # 使用正确的列索引提取金额
                    amount = row[COL_FINAL_AMOUNT] 
                    
                    if subject and amount is not None:
                        all_data.append([year_int, subject, amount])        
        if not all_data:
            logging.warning("未能从业务活动表预制件中收集到任何数据，无法生成收支汇总表。")
            return
        
        full_df = pd.DataFrame(all_data, columns=['年份', '科目', '金额'])
        full_df['金额'] = pd.to_numeric(full_df['金额'], errors='coerce').fillna(0)

        # 3. 创建并注入收入/支出汇总表 (此部分逻辑无需修改)
        _create_pivot_and_inject(wb_to_fill, '收入汇总', full_df[full_df['科目'].isin(income_subjects)], year_header_formatter)
        _create_pivot_and_inject(wb_to_fill, '支出汇总', full_df[full_df['科目'].isin(expense_subjects)], year_header_formatter)

    except (KeyError, IndexError) as e:
        logging.error(f"配置或预制件不完整，无法生成收支汇总表。错误: {e}", exc_info=True)

def _create_pivot_and_inject(wb, sheet_name, df_filtered, year_header_formatter={}):
    """
    【V2 - 签名修复版】
    - 修复了因未同步接收 year_header_formatter 参数导致的 TypeError。
    """
    if df_filtered.empty or sheet_name not in wb.sheetnames: return
    
    ws = wb[sheet_name]
    # 清空旧内容
    # 使用更安全的方式清空，避免 openpyxl 的一些潜在问题
    for row in ws.iter_rows():
        for cell in row: 
            cell.value = None
            
    pivot = pd.pivot_table(df_filtered, values='金额', index='科目', columns='年份', aggfunc='sum').fillna(0)

    # 在写入Excel前，重命名DataFrame的列
    if year_header_formatter:
        rename_map = {year: header for year, header in year_header_formatter.items() if year in pivot.columns}
        pivot.rename(columns=rename_map, inplace=True)
    
    pivot['合计'] = pivot.sum(axis=1)
    pivot.loc['合计'] = pivot.sum()
    pivot = pivot.reset_index()

    # 使用 openpyxl 的 dataframe_to_rows 功能高效注入新内容
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(pivot, index=False, header=True):
        ws.append(r)