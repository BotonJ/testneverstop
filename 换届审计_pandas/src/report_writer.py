# /src/report_writer.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from src.utils.logger_config import logger

class ReportWriter:
    """
    【最终版报告生成器】
    负责将处理和复核后的数据，以专业格式写入最终的Excel报告。
    其核心逻辑完全借鉴并升华了您经典的 inject_modules 和 excel_writer。
    """
    def __init__(self, output_filepath: str):
        self.filepath = output_filepath
        self.workbook = Workbook()
        # 删除默认创建的Sheet，我们将完全自定义
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        logger.info(f"初始化报告写入器，目标文件: {self.filepath}")

    def generate_full_report(self, summary_dict: dict, verification_results: list, normal_df: pd.DataFrame, total_df: pd.DataFrame, mapping: dict):
        """
        生成完整报告的主函数。
        """
        logger.info("--- [注入阶段] 开始生成完整Excel报告 ---")
        
        # 1. 注入业务活动表汇总
        self._write_biz_summary_sheets(normal_df, mapping)
        
        # 2. 注入资产负债表变动表
        # (此部分逻辑较为复杂，可作为后续精修，暂时创建一个占位Sheet)
        self.workbook.create_sheet("资产负债变动")
        logger.info("  -> 已创建'资产负债变 động'占位Sheet页。")
        
        # 3. 注入复核报告和最终说明文字
        self._write_summary_and_verification_sheet("审计说明及复核", summary_dict, verification_results)

        # 4. 保存文件
        self.save()
        logger.info("--- [注入阶段] 完整Excel报告生成完毕 ---")

    def _write_biz_summary_sheets(self, normal_df: pd.DataFrame, mapping: dict):
        """
        借鉴 biz.py 的逻辑，创建收入和支出汇总表。
        """
        logger.info("  -> 正在注入'收入汇总'与'支出汇总'Sheet页...")
        yewu_config = mapping.get("yewu_subtotal_config", {})
        
        income_subjects = yewu_config.get("收入", [])
        expense_subjects = yewu_config.get("费用", [])

        # 准备数据
        income_data = normal_df[normal_df.index.isin(income_subjects)]
        expense_data = normal_df[normal_df.index.isin(expense_subjects)]

        # 创建并写入收入汇总
        if not income_data.empty:
            ws_income = self.workbook.create_sheet("收入汇总")
            income_pivot = pd.pivot_table(income_data.reset_index(), values=income_data.columns, index='项目').reset_index()
            self._write_df_to_sheet(ws_income, income_pivot, "收入汇总")

        # 创建并写入支出汇总
        if not expense_data.empty:
            ws_expense = self.workbook.create_sheet("支出汇总")
            expense_pivot = pd.pivot_table(expense_data.reset_index(), values=expense_data.columns, index='项目').reset_index()
            self._write_df_to_sheet(ws_expense, expense_pivot, "费用汇总")
            
    def _write_summary_and_verification_sheet(self, sheet_name: str, summary: dict, verification: list):
        """
        在一个Sheet中写入最终的JSON汇总和复核报告。
        """
        logger.info(f"  -> 正在注入'{sheet_name}'Sheet页...")
        ws = self.workbook.create_sheet(sheet_name)
        current_row = 1
        
        # 写入JSON汇总
        ws.cell(row=current_row, column=1, value="--- 核心指标汇总 ---")
        current_row += 1
        for key, value in summary.items():
            ws.cell(row=current_row, column=1, value=key)
            ws.cell(row=current_row, column=2, value=str(value)) # 确保所有值都为可写格式
            current_row += 1
        
        current_row += 1 # 空一行
        
        # 写入复核报告
        ws.cell(row=current_row, column=1, value="--- 数据内部复核报告 ---")
        current_row += 1
        for line in verification:
            cell = ws.cell(row=current_row, column=1, value=line)
            if "✅" in line: cell.font = Font(color="008000") # Green
            elif "❌" in line: cell.font = Font(color="FF0000") # Red
            current_row += 1

    def _write_df_to_sheet(self, ws, df, title):
        """通用函数：将DataFrame写入指定的worksheet，并应用基本格式。"""
        # 写入标题
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        
        # 写入DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 应用格式 (简化版)
        self._apply_table_style(ws, 3, ws.max_row, 1, ws.max_column)

    def _apply_table_style(self, ws, start_row, end_row, start_col, end_col):
        """借鉴 excel_writer.py 的通用格式刷。"""
        header_font = Font(name="宋体", size=12, bold=True)
        content_font = Font(name="宋体", size=12)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # 表头格式
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row - 1, column=col_idx)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
        
        # 数据行格式
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = content_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    def save(self):
        """保存工作簿到文件。"""
        try:
            self.workbook.save(self.filepath)
            logger.info(f"✅ Excel报告已成功保存到: {self.filepath}")
        except Exception as e:
            logger.error(f"❌ 错误：保存Excel文件时发生错误: {e}")