# /src/report_generator.py

import logging
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
import pandas as pd
import re
import shutil

# 从新的格式化模块导入函数
from src.report_formatters.format_balance import format_balance_sheet
from src.report_formatters.format_yewu import format_yewu_sheet
from src.report_formatters.format_change_tables import populate_balance_change_sheet
from src.report_formatters.format_biz_summary import create_and_inject_biz_summary

logger = logging.getLogger(__name__)

# --- [新增] ---
# 1. 将您测试代码的核心逻辑封装成一个私有的辅助函数
# 这个函数只在 report_generator.py 内部使用
def _get_dynamic_yewu_headers(year: int, audit_period_str: str) -> tuple[str, str]:
    """
    【V2 - 核心逻辑函数】
    根据给定的年份和审计期间，一次性计算并返回智能格式化的
    (期初表头, 期末表头) 元组。
    此函数已修复跨年期间的期末计算BUG，并能正确处理完整年度的简化。
    """
    # 默认值
    default_qichu = f"{year - 1}年累计数"
    default_qimo = f"{year}年累计数"

    if not audit_period_str or not isinstance(audit_period_str, str):
        return default_qichu, default_qimo

    match = re.match(r'(\d{4})年(\d{1,2})月[-至](\d{4})年(\d{1,2})月', audit_period_str.replace(" ", ""))
    if not match:
        return default_qichu, default_qimo

    start_year, start_month, end_year, end_month = map(int, match.groups())

    header_qichu = default_qichu
    header_qimo = default_qimo

    # --- 开始应用规则 ---
    is_start_year = (year == start_year)
    is_end_year = (year == end_year)
    is_middle_year = (year > start_year and year < end_year)

    # 1. 处理期初 (Qichu)
    if is_start_year and start_month > 1:
        header_qichu = f"{year}年1-{start_month - 1}月累计数"
    
    # 2. 处理期末 (Qimo)
    if is_start_year and is_end_year: # 同一年内
        if start_month > 1 or end_month < 12:
            header_qimo = f"{year}年{start_month}-{end_month}月累计数"
    elif is_start_year: # 跨年期间的起始年
        if start_month > 1:
            header_qimo = f"{year}年{start_month}-12月累计数"
    elif is_end_year: # 跨年期间的终止年
        if end_month < 12:
            header_qimo = f"{year}年1-{end_month}月累计数"
    
    return header_qichu, header_qimo

def _render_header_on_sheet(ws_tgt, year, audit_end_year, header_config):
    """
    【V4 - 最终版】
    调用新的辅助函数 _get_dynamic_yewu_headers 来高效获取期初和期末表头。
    """
    if not header_config: return
    try:
        audit_period_string = header_config.get("期末", {}).get("rule")

        # --- [核心修改] ---
        # 一次性计算出业务活动表所需的期初和期末表头
        yewu_header_qichu, yewu_header_qimo = _get_dynamic_yewu_headers(year, audit_period_string)

        for field_name, meta in header_config.items():
            target_cell, value_to_write = meta.get("target_cell"), None
            if not target_cell: continue

            if "资产负债表" in ws_tgt.title:
                if field_name == "期初": value_to_write = f"{year - 1}年12月31日"
                elif field_name == "期末": value_to_write = f"{year}年12月31日"
                
            elif "业务活动表" in ws_tgt.title:
                # --- [核心修改] ---
                # 直接使用预先计算好的值
                if field_name == "期初": value_to_write = yewu_header_qichu
                elif field_name == "期末": value_to_write = yewu_header_qimo
            
            if field_name == "单位名称":
                value_to_write = meta.get("rule")

            if value_to_write:
                for cell in str(target_cell).split(','):
                    ws_tgt[cell.strip()] = value_to_write

    except Exception as e:
        logger.warning(f"为 Sheet '{ws_tgt.title}' 渲染表头时失败: {e}")

def apply_global_formatting(wb):
    """应用专业的全局数字格式。"""
    logger.info("  -> 应用全局数字格式...")
    summary_format, activity_format = '#,##0.00;-#,##0.00;"-"', '#,##0.00;-#,##0.00;;@'
    right_align = Alignment(horizontal='right', vertical='center')
    sheet_names = ['资产负债变动', '收入汇总', '支出汇总'] + [s.title for s in wb if re.match(r"^\d{4}_(资产负债表|业务活动表)$", s.title)]
    for name in sheet_names:
        if name not in wb.sheetnames: continue
        ws, formatter = wb[name], activity_format if "业务活动表" in name else summary_format
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format, cell.alignment = formatter, right_align

def _create_prebuilt_workbook_in_memory(raw_df: pd.DataFrame, mapping_configs: dict):
    """【V3 - 标准化修复版】根据raw_df在内存中创建数据预制件。"""
    logger.info("  -> 正在内存中创建数据预制件...")
    wb = Workbook()
    wb.remove(wb.active)

    for year, year_df in raw_df.groupby('年份'):
        # --- 创建资产负债表预制Sheet (此部分逻辑不变，已是标准格式) ---
        bs_df = year_df[year_df['报表类型'] == '资产负债表']
        if not bs_df.empty:
            ws_bs = wb.create_sheet(title=f"{year}资产负债表")
            ws_bs.append(["项目", "期初金额", "期末金额"])
            for _, row in bs_df.iterrows():
                ws_bs.append([row['项目'], row['期初金额'], row['期末金额']])
            
            # (调试打印部分保持不变)

        # --- [核心修改] 创建业务活动表预制Sheet ---
        # 现在业务活动表也使用和资产负债表完全相同的、标准的三列表格结构。
        # 这修复了之前只写入期末值、导致期初值丢失的BUG。
        is_df = year_df[year_df['报表类型'] == '业务活动表']
        if not is_df.empty:
            ws_is = wb.create_sheet(title=f"{year}业务活动表")
            # 写入标准表头
            ws_is.append(["项目", "期初金额", "期末金额"])
            # 逐行写入数据
            for _, row in is_df.iterrows():
                # 跳过没有项目的无效行（例如之前自动计算的'净资产变动额'可能没有期初）
                if pd.notna(row['项目']):
                    ws_is.append([row['项目'], row['期初金额'], row['期末金额']])
    return wb

def generate_master_report(raw_df, summary_values, mapping_configs, template_path, output_path):
    """
    【V5.6 - 精确期初逻辑版】
    - 在调用 format_yewu_sheet 时，传入一个新的布尔标志 is_first_audit_year。
    """
    logger.info("--- [报告生成模块启动] ---")

    # 1. 准备报告文件
    try:
        shutil.copy(template_path, output_path)
    except Exception as e:
        logger.error(f"复制模板文件时出错: {e}"); return

    # 2. 在内存中创建数据预制件
    prebuilt_wb = _create_prebuilt_workbook_in_memory(raw_df, mapping_configs)

    # 3. 加载刚刚复制好的报告文件
    wb_final = openpyxl.load_workbook(output_path)
        
    # 4. 创建年度报表
    logger.info("-> 步骤 A: 创建并格式化年度报表 Sheets...")
    alias_dict = mapping_configs.get("alias_dict", {})
    yewu_map_df = mapping_configs.get("业务活动表逐行", pd.DataFrame())
    
    header_meta = mapping_configs.get("HeaderMapping", {})
    header_balance_config = header_meta.get("资产负债表", {})
    header_yewu_config = header_meta.get("业务活动表", {})
    
    years = sorted(raw_df['年份'].unique())
    start_year, end_year = years[0], years[-1]
    prev_yewu_ws = None

    for year in years:
        # --- 处理资产负债表 (此部分无变化) ---
        sheet_name_balance_src = f"{year}资产负债表"
        if sheet_name_balance_src in prebuilt_wb.sheetnames:
            ws_src_balance = prebuilt_wb[sheet_name_balance_src]
            ws_tgt_balance = wb_final.copy_worksheet(wb_final["资产负债表"])
            ws_tgt_balance.title = f"{year}_资产负债表"
            format_balance_sheet(ws_src_balance, ws_tgt_balance, alias_dict)
            _render_header_on_sheet(ws_tgt_balance, year, end_year, header_balance_config)
        
        # --- 处理业务活动表 ---
        sheet_name_yewu_src = f"{year}业务活动表"
        if sheet_name_yewu_src in prebuilt_wb.sheetnames:
            ws_src_yewu = prebuilt_wb[sheet_name_yewu_src]
            net_change = summary_values.get(f"{year}_净资产变动额")
            ws_tgt_yewu = wb_final.copy_worksheet(wb_final["业务活动表"])
            ws_tgt_yewu.title = f"{year}_业务活动表"

            # --- [核心修改] ---
            # 判断当前年份是否为审计期间的第一年
            is_first_year = (year == start_year)
            # 将这个布尔值作为新参数传入
            format_yewu_sheet(
                ws_src_yewu, 
                ws_tgt_yewu, 
                yewu_map_df, 
                prev_ws=prev_yewu_ws, 
                net_asset_change=net_change,
                is_first_audit_year=is_first_year # 新增的标志位
            )
            # --- [核心修改结束] ---

            _render_header_on_sheet(ws_tgt_yewu, year, end_year, header_yewu_config)
            prev_yewu_ws = ws_tgt_yewu

    # 5. 填充分析报表
    logger.info("-> 步骤 B: 填充分析与汇总 Sheets...")
    populate_balance_change_sheet(prebuilt_wb, wb_final, mapping_configs)
    create_and_inject_biz_summary(prebuilt_wb, wb_final, mapping_configs)
    
    # 6. 最终修饰
    logger.info("-> 步骤 C: 应用全局格式化并清理...")
    apply_global_formatting(wb_final)
    if '资产负债表' in wb_final.sheetnames: wb_final.remove(wb_final['资产负债表'])
    if '业务活动表' in wb_final.sheetnames: wb_final.remove(wb_final['业务活动表'])
    
    # 7. 保存
    try:
        wb_final.save(output_path)
        logger.info(f"✅✅✅ 最终审计报告已成功生成: {output_path}")
    except Exception as e:
        logger.error(f"保存最终报告时失败: {e}", exc_info=True) 