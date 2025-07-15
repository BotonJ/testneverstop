# /src/classic_report_generator.py
import re
import pandas as pd
from openpyxl import load_workbook
from src.utils.logger_config import logger
from modules.utils import normalize_name

class ClassicReportGenerator:
    """
    【最终修复版 V2】
    增加了更强大的智能Sheet名匹配逻辑。
    """
    def __init__(self, source_path, template_path, mapping):
        self.wb_src = load_workbook(source_path, data_only=True)
        self.wb_tgt = load_workbook(template_path)
        self.mapping = mapping
        # 构建一个【清洗过的】源Sheet名查找字典
        self.src_sheet_map = {normalize_name(name): name for name in self.wb_src.sheetnames}
        self.alias_lookup = self._build_alias_lookup()
        logger.info("经典报告生成器已初始化。")

    def _build_alias_lookup(self):
        # ... (此函数保持不变) ...
        alias_lookup = {}
        alias_map_df = self.mapping.get("alias_map_df")
        if alias_map_df is not None:
            for _, row in alias_map_df.iterrows():
                standard = normalize_name(row['标准科目名'])
                if not standard: continue
                for col in alias_map_df.columns:
                    if '等价科目名' in col and pd.notna(row[col]):
                        aliases = [normalize_name(alias) for alias in str(row[col]).split(',')]
                        for alias in aliases:
                            if alias: alias_lookup[alias] = standard
        return alias_lookup

    def _fill_balance_sheet(self, ws_src, ws_tgt):
        # ... (此函数保持不变) ...
        logger.info(f"  -> 正在向'{ws_tgt.title}'注入资产负债表数据...")
        src_dict = {}
        for i in range(1, ws_src.max_row + 1):
            name_a_raw, name_e_raw = ws_src[f"A{i}"].value, ws_src[f"E{i}"].value
            if name_a_raw:
                name_a_clean = normalize_name(name_a_raw)
                if name_a_clean:
                    name_std = self.alias_lookup.get(name_a_clean, name_a_clean)
                    src_dict[name_std] = {"期初": ws_src[f"C{i}"].value, "期末": ws_src[f"D{i}"].value}
            if name_e_raw:
                name_e_clean = normalize_name(name_e_raw)
                if name_e_clean:
                    name_std = self.alias_lookup.get(name_e_clean, name_e_clean)
                    if name_std not in src_dict:
                        src_dict[name_std] = {"期初": ws_src[f"G{i}"].value, "期末": ws_src[f"H{i}"].value}
        for row in ws_tgt.iter_rows(min_row=2):
            tgt_cell = row[0]
            if tgt_cell.value:
                tgt_name_clean = normalize_name(tgt_cell.value)
                if tgt_name_clean in src_dict:
                    ws_tgt.cell(row=tgt_cell.row, column=2, value=src_dict[tgt_name_clean]["期初"])
                    ws_tgt.cell(row=tgt_cell.row, column=3, value=src_dict[tgt_name_clean]["期末"])

    def _fill_income_statement(self, ws_src, ws_tgt):
        # ... (此函数保持不变) ...
        logger.info(f"  -> 正在向'{ws_tgt.title}'注入业务活动表数据...")
        yewu_line_map = self.mapping.get("yewu_line_map", [])
        for item in yewu_line_map:
            src_initial, src_final = item.get("源期初坐标"), item.get("源期末坐标")
            tgt_initial, tgt_final = item.get("目标期初坐标"), item.get("目标期末坐标")
            if pd.notna(src_initial) and pd.notna(tgt_initial):
                try: ws_tgt[tgt_initial].value = ws_src[src_initial].value
                except: pass
            if pd.notna(src_final) and pd.notna(tgt_final):
                try: ws_tgt[tgt_final].value = ws_src[src_final].value
                except: pass

    def _find_matching_src_sheet(self, target_sheet_name):
        """【最终修复版】智能查找函数"""
        target_clean = normalize_name(target_sheet_name)
        
        # 1. 尝试直接匹配清洗过的名称
        if target_clean in self.src_sheet_map:
            return self.src_sheet_map[target_clean]

        # 2. 尝试模式匹配
        match = re.search(r'(\d{4})', target_clean)
        if not match: return None
        
        year = match.group(1)
        # 构建可能的简化版名称并清洗
        possible_names = []
        if "资产负债表" in target_clean:
            possible_names.append(normalize_name(f"{year}z"))
        elif "业务活动表" in target_clean:
            possible_names.append(normalize_name(f"{year}y"))

        # 在我们清洗过的源Sheet名查找字典中寻找
        for name in possible_names:
            if name in self.src_sheet_map:
                return self.src_sheet_map[name]
        
        return None

    def create_report(self, output_path):
        logger.info("--- [经典流程移植] 开始生成格式化的'新soce'... ---")
        
        for ws_tgt in self.wb_tgt:
            target_sheet_name = ws_tgt.title
            matching_src_name = self._find_matching_src_sheet(target_sheet_name)
            
            if matching_src_name:
                logger.info(f"成功匹配模板Sheet '{target_sheet_name}' -> 源Sheet '{matching_src_name}'")
                ws_src = self.wb_src[matching_src_name]
                
                if "资产负债表" in target_sheet_name:
                    self._fill_balance_sheet(ws_src, ws_tgt)
                elif "业务活动表" in target_sheet_name:
                    self._fill_income_statement(ws_src, ws_tgt)
            else:
                 logger.warning(f"在源文件中未找到与模板Sheet '{target_sheet_name}' 匹配的任何源Sheet。")

        try:
            self.wb_tgt.save(output_path)
            logger.info(f"✅ '新soce'已成功生成到: {output_path}")
        except Exception as e:
            logger.error(f"❌ 保存'新soce'时发生错误: {e}")