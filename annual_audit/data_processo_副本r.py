import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

class DataProcessor:
    """
    【最终版本】核心数据处理器：使用openpyxl进行精确数据提取，再交由pandas进行处理。
    """
    def __init__(self, source_filepath: str, configs_dict: dict):
        self.source_filepath = source_filepath
        self.configs = configs_dict
        self.raw_extracted_data = []
        self.processed_data = {}
        self.verification_totals = {}
        print("初始化数据处理器 (最终版本)。")

    def _get_column_index(self, col_str: str) -> int:
        return column_index_from_string(str(col_str))

    def _extract_verification_totals(self):
        """
        【最终修复版】一个专门的函数，用于提取所有复核所需的总计值。
        此函数完全由 mapping_file 驱动，健壮且能正确处理多栏布局和别名。
        """
        print("  正在专门提取用于复核的总计值...")
        try:
            wb = load_workbook(self.source_filepath, data_only=True)
            alias_map_df = self.configs.get('科目等价映射', pd.DataFrame())
            bs_map = self.configs.get('资产负债表区块', pd.DataFrame())
            act_map = self.configs.get('业务活动表逐行', pd.DataFrame())

            # 1. 创建别名 -> 标准名的反向映射字典
            alias_to_standard_map = {}
            if not alias_map_df.empty:
                for _, row in alias_map_df.iterrows():
                    standard_name = row['标准科目名']
                    # 确保标准名本身也作为自己的“别名”
                    alias_to_standard_map[standard_name] = standard_name
                    for col in alias_map_df.columns:
                        if '等价科目名' in col and pd.notna(row[col]):
                            aliases = [alias.strip() for alias in str(row[col]).split(',')]
                            for alias in aliases:
                                if alias: alias_to_standard_map[alias] = standard_name

            # 2. 定义我们需要查找的所有总计项的“指令清单”
            target_totals_config = {
                '净资产合计': ['期初净资产', '期末净资产', bs_map, '区块名称', wb['资产负债表']],
                '收入合计': ['收入合计', None, act_map, '字段名', wb['业务活动表']],
                '费用合计': ['费用合计', None, act_map, '字段名', wb['业务活动表']],
            }

            # 3. 遍历“指令清单”，统一处理每一个要查找的项目
            for std_name, config in target_totals_config.items():
                start_key, end_key, map_df, map_item_col, sheet = config
                
                aliases_to_find = [alias for alias, std in alias_to_standard_map.items() if std == std_name]
                if not aliases_to_find: aliases_to_find.append(std_name)

                found = False
                for row_idx in range(1, 61):
                    cells_to_check = [sheet.cell(row=row_idx, column=1)]
                    if sheet.title == '资产负债表':
                        cells_to_check.append(sheet.cell(row=row_idx, column=5))

                    for cell in cells_to_check:
                        if not cell.value: continue
                        cell_text_clean = str(cell.value).strip()

                        for alias in aliases_to_find:
                            if alias == cell_text_clean:
                                print(f"    -> 在 '{sheet.title}' 第 {cell.row} 行根据别名 '{alias}' 命中 '{std_name}'")
                                
                                config_row = map_df[map_df[map_item_col] == std_name]
                                if config_row.empty:
                                    print(f"      ⚠️ 警告: 在mapping文件中找不到标准名 '{std_name}' 的配置。")
                                    continue

                                if sheet.title == '资产负债表':
                                    start_col = config_row['期初列'].iloc[0]
                                    end_col = config_row['期末列'].iloc[0]
                                    start_val = sheet.cell(row=cell.row, column=self._get_column_index(start_col)).value
                                    end_val = sheet.cell(row=cell.row, column=self._get_column_index(end_col)).value
                                    
                                    self.verification_totals[start_key] = pd.to_numeric(start_val, errors='coerce')
                                    if end_key: self.verification_totals[end_key] = pd.to_numeric(end_val, errors='coerce')
                                    print(f"      -> 已提取: {start_key}={self.verification_totals.get(start_key, 'N/A')}, {end_key}={self.verification_totals.get(end_key, 'N/A')}")
                                else:
                                    end_col = config_row['期末合计列'].iloc[0]
                                    end_val = sheet.cell(row=cell.row, column=self._get_column_index(end_col)).value
                                    self.verification_totals[start_key] = pd.to_numeric(end_val, errors='coerce')
                                    print(f"      -> 已提取: {start_key}={self.verification_totals.get(start_key, 'N/A')}")
                                
                                found = True
                                break
                        if found: break
                    if found: break
                
                if not found:
                    print(f"  ⚠️ 警告: 未能在文件中找到任何与 '{std_name}' 匹配的合计项。")

            print("\n  --- 复核所需总计值提取完毕 ---")
            print(f"  {self.verification_totals}")

        except Exception as e:
            print(f"❌ 错误: 在提取复核总计值时发生异常: {e}")

    def _parse_balance_sheet(self):
        # 此函数的内部逻辑保持不变
        print("  正在解析'资产负债表'...")
        wb = load_workbook(self.source_filepath, data_only=True)
        sheet = wb['资产负债表']
        bs_map = self.configs.get('资产负债表区块', pd.DataFrame())
        if bs_map.empty: return

        for _, row_map in bs_map.iterrows():
            start_cell, end_cell = row_map['起始单元格'], row_map['终止单元格']
            start_row, end_row = int(re.search(r'\d+', start_cell).group()), int(re.search(r'\d+', end_cell).group())
            item_col_idx = self._get_column_index(re.search(r'[A-Z]+', start_cell).group())
            start_val_col_idx, end_val_col_idx = self._get_column_index(row_map['期初列']), self._get_column_index(row_map['期末列'])
            
            skip_keywords = str(row_map['跳过行']).split(',') if pd.notna(row_map['跳过行']) else []
            group_name = row_map.get('附注组名')
            is_note_item = row_map.get('是否为附注科目') == '是'

            for row_idx in range(start_row, end_row + 1):
                item_name = sheet.cell(row=row_idx, column=item_col_idx).value
                if not item_name or not isinstance(item_name, str) or item_name.isspace(): continue
                item_name = item_name.strip()
                if any(keyword in item_name for keyword in skip_keywords if keyword): continue

                start_val = sheet.cell(row=row_idx, column=start_val_col_idx).value
                end_val = sheet.cell(row=row_idx, column=end_val_col_idx).value
                
                if item_name:
                    self.raw_extracted_data.append({
                        "项目": item_name, "期初数": start_val, "期末数": end_val,
                        "来源表": "资产负债表", 
                        "附注组名": group_name if pd.notna(group_name) else item_name,
                        "是否为附注科目": is_note_item
                    })
        print("  '资产负债表'解析完成。")

    def _parse_activity_sheet(self):
        # 此函数的内部逻辑保持不变
        print("  正在解析'业务活动表'...")
        wb = load_workbook(self.source_filepath, data_only=True)
        sheet = wb['业务活动表']
        act_map = self.configs.get('业务活动表逐行', pd.DataFrame())
        if act_map.empty: return

        all_items_map = act_map.to_dict('records')
        row_offset = 0

        for row_map in all_items_map:
            item_name_map = str(row_map['字段名']).strip()
            group_name = row_map.get('附注组名')
            is_note_item = row_map.get('是否为附注科目') == '是'
            
            if pd.notna(row_map['行号']):
                row_num_map = int(row_map['行号'])
                is_income = '收入' in str(group_name)
                
                if "商品销售收入" in item_name_map:
                    actual_item_name = sheet.cell(row=row_num_map, column=1).value
                    if actual_item_name and item_name_map in str(actual_item_name):
                        end_val_cell = sheet.cell(row=row_num_map, column=self._get_column_index(row_map['期末合计列']))
                        if pd.notna(end_val_cell.value) and float(end_val_cell.value) != 0:
                            row_offset = 1
                
                target_row = row_num_map + (row_offset if is_income else 0)
                actual_item_name_check = sheet.cell(row=target_row, column=1).value
                actual_clean = str(actual_item_name_check).strip().replace('　', '')

                if actual_item_name_check and item_name_map == actual_clean:
                    start_val = sheet.cell(row=target_row, column=self._get_column_index(row_map['期初合计列'])).value
                    end_val = sheet.cell(row=target_row, column=self._get_column_index(row_map['期末合计列'])).value
                    self.raw_extracted_data.append({
                        "项目": item_name_map, "期初数": start_val, "期末数": end_val,
                        "来源表": "业务活动表", "附注组名": group_name if pd.notna(group_name) else item_name_map,
                        "是否为附注科目": is_note_item
                    })
            else:
                if '费用' in item_name_map or '成本' in item_name_map:
                    found = False
                    for row in sheet.iter_rows(min_row=10, max_row=50, min_col=1, max_col=1):
                        cell = row[0]
                        if cell.value and item_name_map in str(cell.value):
                            start_val = sheet.cell(row=cell.row, column=self._get_column_index(row_map['期初合计列'])).value
                            end_val = sheet.cell(row=cell.row, column=self._get_column_index(row_map['期末合计列'])).value
                            self.raw_extracted_data.append({
                                "项目": item_name_map, "期初数": start_val, "期末数": end_val,
                                "来源表": "业务活动表", "附注组名": group_name if pd.notna(group_name) else item_name_map,
                                "是否为附注科目": is_note_item
                            })
                            found = True
                            break
        print("  '业务活动表'解析完成。")

    def get_notes_data(self) -> pd.DataFrame:
        # 主调用流程保持不变
        print("正在处理'报表附注'数据...")
        self.raw_extracted_data = []
        self._extract_verification_totals()
        self._parse_balance_sheet()
        self._parse_activity_sheet()
        
        if not self.raw_extracted_data: return pd.DataFrame()
        df = pd.DataFrame(self.raw_extracted_data)
        df = df[df['是否为附注科目'] == True].copy()
        df['期末数'] = pd.to_numeric(df['期末数'], errors='coerce').fillna(0)
        df['期初数'] = pd.to_numeric(df['期初数'], errors='coerce').fillna(0)
        
        net_asset_group_name = '净资产'
        net_asset_df = df[df['附注组名'] == net_asset_group_name].copy()
        other_items_df = df[df['附注组名'] != net_asset_group_name].copy()
        other_items_filtered_df = other_items_df[(other_items_df['期末数'].abs() > 1e-6) | (other_items_df['期初数'].abs() > 1e-6)].copy()
        
        bs_items_df = other_items_filtered_df[other_items_filtered_df['来源表'] == '资产负债表']
        act_items_df = other_items_filtered_df[other_items_filtered_df['来源表'] == '业务活动表']
        
        final_df = pd.concat([bs_items_df, net_asset_df, act_items_df], ignore_index=True)
        final_df['项目'] = final_df['项目'].apply(lambda x: re.sub(r'^[（\(][一二三四五六七八九十\d]+[）\)]\s*', '', str(x)).strip())
        final_df['附注组名'] = final_df['附注组名'].apply(lambda x: re.sub(r'^[（\(][一二三四五六七八九十\d]+[）\)]\s*', '', str(x)).strip())
        final_df.reset_index(drop=True, inplace=True)
        
        self.processed_data['notes_data'] = final_df
        return final_df

    def run_verification_checks(self) -> list:
        # 此函数保持不变
        print("--- 开始执行数据复核 ---")
        results = []
        notes_df = self.processed_data.get('notes_data', pd.DataFrame())
        if notes_df.empty:
            results.append("未提取到有效数据，无法执行复核。")
            return results

        income_group_name = '收入'
        calc_income_total = notes_df[notes_df['附注组名'] == income_group_name]['期末数'].sum()
        report_income_total = self.verification_totals.get('收入合计', 0)
        if abs(calc_income_total - report_income_total) < 1e-6:
            results.append(f"✅ 收入内部核对成功: 计算值 {calc_income_total:,.2f} vs 报表值 {report_income_total:,.2f}")
        else:
            diff = calc_income_total - report_income_total
            results.append(f"❌ 收入内部核对失败: 计算值 {calc_income_total:,.2f} vs 报表值 {report_income_total:,.2f} (差额: {diff:,.2f})")

        expense_items = ['业务活动成本', '管理费用', '筹资费用', '其他费用']
        calc_expense_total = notes_df[notes_df['项目'].isin(expense_items)]['期末数'].sum()
        report_expense_total = self.verification_totals.get('费用合计', 0)
        if abs(calc_expense_total - report_expense_total) < 1e-6:
            results.append(f"✅ 支出内部核对成功: 计算值 {calc_expense_total:,.2f} vs 报表值 {report_expense_total:,.2f}")
        else:
            diff = calc_expense_total - report_expense_total
            results.append(f"❌ 支出内部核对失败: 计算值 {calc_expense_total:,.2f} vs 报表值 {report_expense_total:,.2f} (差额: {diff:,.2f})")

        income_minus_expense = report_income_total - report_expense_total
        net_asset_change = self.verification_totals.get('期末净资产', 0) - self.verification_totals.get('期初净资产', 0)
        if abs(income_minus_expense - net_asset_change) < 1e-6:
            results.append(f"✅ 收支与净资产联动核对成功: 收支差额 {income_minus_expense:,.2f} vs 净资产变动 {net_asset_change:,.2f}")
        else:
            diff = income_minus_expense - net_asset_change
            results.append(f"❌ 收支与净资产联动核对失败: 收支差额 {income_minus_expense:,.2f} vs 净资产变动 {net_asset_change:,.2f} (差额: {diff:,.2f})")
        
        print("--- 复核结束 ---")
        return results

    def extract_audit_year(self) -> int | None:
        # 此函数保持不变
        print("正在自动提取审计年度...")
        try:
            wb = load_workbook(self.source_filepath, data_only=True)
            bs_sheet, act_sheet = wb['资产负债表'], wb['业务活动表']
            pattern_date = re.compile(r'(\d{4})年12月31日')
            pattern_year = re.compile(r'(\d{4})年度')
            bs_year, act_year = None, None
            
            for cell in bs_sheet[3]:
                if cell.value:
                    if isinstance(cell.value, str):
                        match = pattern_date.search(cell.value)
                        if match: bs_year = int(match.group(1)); break
                    elif hasattr(cell.value, 'year'):
                        bs_year = cell.value.year; break
            
            for cell in act_sheet[3]:
                if cell.value and isinstance(cell.value, str):
                    match = pattern_year.search(cell.value)
                    if match: act_year = int(match.group(1)); break

            if bs_year and act_year and bs_year == act_year:
                print(f"✅ 审计年度验证成功: {bs_year}")
                return bs_year
            else:
                print(f"❌ 错误：未能从两张表中找到一致的审计年度 (资负: {bs_year}, 业务: {act_year})。")
                return None
        except Exception as e:
            print(f"❌ 错误：在提取审计年度时发生异常: {e}")
            return None