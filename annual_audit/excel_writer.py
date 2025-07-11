import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelWriter:
    def __init__(self, output_filepath: str):
        self.filepath = output_filepath
        self.workbook = Workbook()
        self.default_sheet = self.workbook.active
        print(f"初始化Excel写入器，目标文件: {self.filepath}")

    def write_notes_sheet(self, sheet_name: str, intro_text: str, notes_df: pd.DataFrame, verification_report: list):
        """
        【最终修复版】在一个新的Sheet页中，写入“报表附注”的所有内容。
        严格遵循正确的写入顺序：1.引言 -> 2.复核报告 -> 3.附注表格。
        """
        print(f"正在创建并写入 '{sheet_name}' Sheet页...")

        if sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            ws.delete_rows(1, ws.max_row + 1)
        else:
            ws = self.workbook.create_sheet(title=sheet_name)

        # 定义样式
        intro_font = Font(name="宋体", size=12)
        group_title_font = Font(name="黑体", size=12, bold=True)
        verification_font = Font(name="宋体", size=12, italic=True, color="808080")

        current_row = 1

        # --- 步骤 1: 写入引言文本 ---
        if intro_text:
            ws.cell(row=current_row, column=1, value=intro_text).font = intro_font
            ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 2

        # --- 步骤 2: 【核心修复】先写入复核报告 ---
        if verification_report:
            ws.cell(row=current_row, column=1, value="--- 数据内部复核结果 ---").font = Font(name="宋体", size=12, bold=True)
            current_row += 1
            for report_line in verification_report:
                cell = ws.cell(row=current_row, column=1, value=report_line)
                if "✅" in report_line:
                    cell.font = Font(name="宋体", size=12, color="008000") # 绿色
                elif "❌" in report_line:
                    cell.font = Font(name="宋体", size=12, color="FF0000")
                else:
                    cell.font = verification_font
                current_row += 1
            current_row += 1 # 复核报告后空一行

        # --- 步骤 3: 按“附注组名”分组，写入每个表格 ---
        if not notes_df.empty:
            note_number = 1
            for group_name, group_df in notes_df.groupby('附注组名',sort=False):
                ws.cell(row=current_row, column=1, value=f"{note_number}.{group_name}").font = group_title_font                
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1
                note_number += 1

                table_df = group_df[['项目', '期初数', '期末数']].copy()
                total_start = table_df['期初数'].sum()
                total_end = table_df['期末数'].sum()
                total_row = pd.DataFrame([{'项目': '合    计', '期初数': total_start, '期末数': total_end}])
                table_df = pd.concat([table_df, total_row], ignore_index=True)
                for r_idx, record in enumerate(dataframe_to_rows(table_df, index=False, header=True), current_row):
                    for c_idx, value in enumerate(record, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 调用我们统一的格式刷
                self._apply_table_style(ws, current_row, current_row + len(table_df), 1, len(table_df.columns))
                
                current_row += len(table_df) + 2

        # --- 步骤 4: 调整列宽 (使用我们已修复好的版本) ---
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = col[0].column_letter
            start_row_for_calc = 3 if column_letter == 'A' else 1
            for row_idx in range(start_row_for_calc, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell_len = 0
                    if cell.value:
                        for char in str(cell.value):
                            cell_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            adjusted_width = max((min(max_length, 40), 12)) + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        print(f"✅ '{sheet_name}' Sheet页已成功写入并格式化。")
    
    def write_audit_sheet(self, sheet_name: str, tables_dict: dict):
            """
            在一个新的Sheet页中，写入“审计事项说明”的多个表格及其标题。
            """
            print(f"正在创建并写入 '{sheet_name}' Sheet页...")

            if sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                ws.delete_rows(1, ws.max_row + 1)
            else:
                ws = self.workbook.create_sheet(title=sheet_name)

            # 定义标题样式
            title_font = Font(name="黑体", size=12, bold=True)
            current_row = 1

            for title, df in tables_dict.items():
                if df.empty:
                    continue
                # 1. 写入主标题
                ws.cell(row=current_row, column=1, value=title).font = title_font
                current_row += 1

                # 2. 写入DataFrame到Excel
                for r_idx, record in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
                    for c_idx, value in enumerate(record, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 3. 【核心修改】调用通用的样式函数来格式化刚刚写入的表格
                table_start_row = current_row
                table_end_row = current_row + len(df)
                table_start_col = 1
                table_end_col = len(df.columns)
                self._apply_table_style(ws, table_start_row, table_end_row, table_start_col, table_end_col)
                
                # 更新下一张表的起始行
                current_row += len(df) + 2 

            # 自动调整列宽 (逻辑不变)
            # ... (此处的列宽调整代码保持不变) ...
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = 0
                        if cell.value:
                            for char in str(cell.value):
                                cell_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width


            print(f"✅ '{sheet_name}' Sheet页已成功写入。")    

    def _apply_table_style(self, ws, start_row, end_row, start_col, end_col):
        """
        一个通用的函数，为指定的表格范围应用统一的、专业的财务报表样式。
        """
        # 定义字体
        content_font = Font(name="宋体", size=12)
        # 【核心修改】表头字体不再加粗
        header_font = Font(name="宋体", size=12, bold=False)
        
        # 定义边框
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 格式化表头
        header_row_index = start_row
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row_index, column=col_idx)
            cell.font = header_font
            # 【核心修改】背景色填充的逻辑被移除
            # cell.fill = header_fill (这行被删除)
            cell.border = thin_border
            # 【核心修改】表头也居中
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 格式化数据行
        for row_idx in range(start_row + 1, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = content_font
                cell.border = thin_border
                # 【核心修改】所有单元格内容，无论文本还是数字，全部居中
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # 数字格式化逻辑可以保留
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    def save(self):
        # (此函数保持不变)
        try:
            # if self.default_sheet in self.workbook.sheetnames:
            #     self.workbook.remove(self.default_sheet)
            # self.workbook.save(self.filepath)
            if "Sheet" in self.workbook.sheetnames:
                default_sheet = self.workbook["Sheet"]
                # 确保我们创建的Sheet页数量大于1，才删除默认页，避免删掉唯一的Sheet
                if len(self.workbook.sheetnames) > 1:
                    self.workbook.remove(default_sheet)
                    print("  -> 已自动删除默认创建的空白'Sheet'页。")
            self.workbook.save(self.filepath)
            print(f"✅ Excel报告已成功保存到: {self.filepath}")
        except Exception as e:
            print(f"❌ 错误：保存Excel文件时发生错误: {e}")