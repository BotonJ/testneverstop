# inject_modules/table3.py
import logging
from pathlib import Path
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


# 复制这个完整的函数
def _get_top_left_merged_cell_address(ws, a1_address_str):
    """
    Given a worksheet and an A1-style cell address string,
    returns the A1-style address of the top-left cell of the merged region it belongs to.
    If the cell is not merged, returns the original address.
    """
    if not isinstance(a1_address_str, str) or not a1_address_str:
        return a1_address_str

    try:
        row, col = coordinate_to_tuple(a1_address_str)
        target_coord = (row, col)
    except Exception as e:
        return a1_address_str

    for merged_range in ws.merged_cells.ranges:
        if target_coord[0] >= merged_range.min_row and target_coord[0] <= merged_range.max_row and \
           target_coord[1] >= merged_range.min_col and target_coord[1] <= merged_range.max_col:
            
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_address = f"{get_column_letter(min_col)}{min_row}"
            return top_left_address

    return a1_address_str

def _get_value_from_cell(ws: Worksheet, address: str) -> float:
    """安全地从指定单元格获取数值，如果为空则返回0。"""
    if not address or not isinstance(address, str):
        return 0.0
    cell_value = ws[address].value
    return float(cell_value) if cell_value is not None else 0.0

def _apply_formulas_from_mapping(ws_tgt: Worksheet, mapping_file_path: str):
    """
    读取“合计公式配置”Sheet，并根据其内容向目标Sheet注入公式。
    """    
    try:
        # 读取配置，并假设第一行是表头
        df_formula = pd.read_excel(mapping_file_path, sheet_name="合计公式配置")       

        for _, row in df_formula.iterrows():
            target_cell_address = row.get("变动单元格")
            formula = row.get("变动公式")

            if pd.notna(target_cell_address) and pd.notna(formula):
                cell = ws_tgt[target_cell_address]
                cell.value = f'={formula}' # 确保写入的是公式
                cell.number_format = '#,##0.00'
                
    except Exception as e:
        return


def inject_table3(wb_src: Workbook, ws_tgt: Worksheet, conf: dict, df_map, mapping_file_path:str, log=None):
    """
    根据inj3的配置，注入限定性/非限定性净资产数据，并调用公式注入模块。
    """
    start_sheet_name = conf.get("start_sheet")
    end_sheet_name = conf.get("end_sheet")

    if not (start_sheet_name and end_sheet_name and start_sheet_name in wb_src.sheetnames and end_sheet_name in wb_src.sheetnames):
        logging.error("table3配置或源文件Sheet不完整。")
        return

    ws_start = wb_src[start_sheet_name]
    ws_end = wb_src[end_sheet_name]    
    logging.info("进入 inject_table3 函数")

    # --- 第一步：像以前一样，注入所有期初、期末和增减数据 ---
    for _, row in df_map.iterrows():
        # (此部分注入数据的逻辑保持不变，但为了完整性在此列出)
        source_field = row.get("来源字段")
        if not source_field: continue

        # 从源工作表的特定单元格读取值
        val_start = _get_value_from_cell(ws_start, row.get("来源单元格（期初）"))
        val_end = _get_value_from_cell(ws_end, row.get("来源单元格（期末）"))
        change = val_end - val_start

        # 注入期初、期末值
        original_addr_start = row.get("目标单元格（期初）")
        if original_addr_start:
            actual_addr_start = _get_top_left_merged_cell_address(ws_tgt, original_addr_start)
            cell_start = ws_tgt[actual_addr_start]
            cell_start.value = val_start
            cell_start.number_format = '#,##0.00'

        original_addr_end = row.get("目标单元格（期末）")
        if original_addr_end:
            actual_addr_end = _get_top_left_merged_cell_address(ws_tgt, original_addr_end)
            cell_end = ws_tgt[actual_addr_end]
            cell_end.value = val_end
            cell_end.number_format = '#,##0.00'

        # 根据正负，注入到“增加”或“减少”单元格
        if change > 0:
            original_addr_increase = row.get("增加单元格")
            if original_addr_increase: 
                actual_addr_increase = _get_top_left_merged_cell_address(ws_tgt, original_addr_increase)
                ws_tgt[actual_addr_increase].value = change
                ws_tgt[actual_addr_increase].number_format = '#,##0.00'
        elif change < 0:
            original_addr_decrease = row.get("减少单元格")
            if original_addr_decrease:
                actual_addr_decrease = _get_top_left_merged_cell_address(ws_tgt, original_addr_decrease)
                ws_tgt[actual_addr_decrease].value = abs(change) # 写入绝对值
                ws_tgt[actual_addr_decrease].number_format = '#,##0.00'
    # --- 调用公式注入模块 ---
    logging.info("开始从'合计公式配置'注入求和公式...")
    _apply_formulas_from_mapping(ws_tgt, mapping_file_path)

    logging.info("inject_table3 函数执行完毕。")
