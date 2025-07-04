import openpyxl
from openpyxl.utils import get_column_letter # 移除 coordinate_to_tuple 导入

def get_balance_core_data(ws_balance, block_map, alias_dict):
    # print(f"DEBUG_BALANCE: 进入 get_balance_core_data 函数...")
    # print(f"DEBUG_BALANCE: Source Sheet Name: {ws_balance.title}")
    # print(f"DEBUG_BALANCE: Block Map Keys: {list(block_map.keys())}")
    # print(f"DEBUG_BALANCE: Alias Dict Keys (if applicable): {list(alias_dict.keys()) if isinstance(alias_dict, dict) else 'Not a dict'}")

    result = {
        "期初资产总额": 0.0,
        "期末资产总额": 0.0,
        "期初负债总额": 0.0,
        "期末负债总额": 0.0,
        "期初净资产总额": 0.0,
        "期末净资产总额": 0.0,
    }

    # Helper function to safely convert value to float, default to 0.0
    def non_numeric_to_zero(value, coordinate_str="N/A"):
        if isinstance(value, (int, float)):
            return value
        elif isinstance(value, str):
            try:
                # 尝试去除逗号并转换为浮点数，处理可能的空字符串
                cleaned_value = value.replace(",", "").strip()
                if cleaned_value == "":
                    return 0.0
                return float(cleaned_value)
            except ValueError:
                print(f"DEBUG_BALANCE: 警告: 单元格 {coordinate_str} 的值 '{value}' 无法转换为数字，设为 0.0。")
                return 0.0
        else:
            # 对于 NoneType 或其他非数字/字符串类型
            if value is None:
                return 0.0
            #print(f"DEBUG_BALANCE: 警告: 单元格 {coordinate_str} 的值 '{value}' 类型非数字或字符串，设为 0.0。")
            return 0.0

    fields_to_extract = ["资产总额", "负债总额", "净资产总额"] # 这些是您想提取的别名

    for field_alias in fields_to_extract:
        # 使用 alias_dict 获取标准科目名
        std_key = alias_dict.get(field_alias, field_alias)
        #print(f"\nDEBUG_BALANCE: --- 处理字段: '{field_alias}' (标准名: '{std_key}') ---")

        # 从 block_map 中获取区块信息
        block_info = block_map.get(std_key)
        if not block_info:
            #print(f"DEBUG_BALANCE: 警告: 标准科目名 '{std_key}' 未在 block_map 中找到对应信息。跳过。")
            continue

        #print(f"DEBUG_BALANCE: '{std_key}' 的 Block Info: {block_info}")

        # **修正：直接使用 'target_row' 作为目标行号**
        row_to_read = block_info.get("target_row")
        col_initial = block_info.get("target_col_initial") # 目标期初列
        col_final = block_info.get("target_col_final")     # 目标期末列

        # 检查坐标是否缺失
        if row_to_read is None or col_initial is None or col_final is None:
            #print(f"DEBUG_BALANCE: 警告: '{std_key}' 的坐标信息不完整。行号: {row_to_read}, 期初列: {col_initial}, 期末列: {col_final}. 跳过此字段的读取。")
            result[f"期初{field_alias}"] = 0.0
            result[f"期末{field_alias}"] = 0.0
            continue # 移动到下一个字段

        # 读取“期初”值
        initial_cell_coordinate = f"{get_column_letter(col_initial)}{row_to_read}"
        ##print(f"DEBUG_BALANCE: 尝试读取 期初值 从单元格: '{initial_cell_coordinate}'")
        try:
            initial_value_raw = ws_balance[initial_cell_coordinate].value
            initial_value = non_numeric_to_zero(initial_value_raw, initial_cell_coordinate)
            result[f"期初{field_alias}"] = initial_value
            ##print(f"DEBUG_BALANCE: 期初{field_alias} (从 {initial_cell_coordinate}) 读取值: {initial_value_raw} -> {initial_value}")
        except Exception as e:
            print(f"DEBUG_BALANCE: 错误: 读取 期初{field_alias} 单元格 '{initial_cell_coordinate}' 失败: {e}。设为 0.0。")
            result[f"期初{field_alias}"] = 0.0

        # 读取“期末”值
        final_cell_coordinate = f"{get_column_letter(col_final)}{row_to_read}"
        ##print(f"DEBUG_BALANCE: 尝试读取 期末值 从单元格: '{final_cell_coordinate}'")
        try:
            final_value_raw = ws_balance[final_cell_coordinate].value
            final_value = non_numeric_to_zero(final_value_raw, final_cell_coordinate)
            result[f"期末{field_alias}"] = final_value
            ##print(f"DEBUG_BALANCE: 期末{field_alias} (从 {final_cell_coordinate}) 读取值: {final_value_raw} -> {final_value}")
        except Exception as e:
            ##print(f"DEBUG_BALANCE: 错误: 读取 期末{field_alias} 单元格 '{final_cell_coordinate}' 失败: {e}。设为 0.0。")
            result[f"期末{field_alias}"] = 0.0
            
    ##print(f"\nDEBUG_BALANCE: 所有核心余额数据读取完毕。最终结果: {result}")
    return result