# File: inject_modules/table1.py
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

def _get_top_left_merged_cell_address(ws, a1_address_str):
    """
    Given a worksheet and an A1-style cell address string,
    returns the A1-style address of the top-left cell of the merged region it belongs to.
    If the cell is not merged, returns the original address.
    """
    if not isinstance(a1_address_str, str) or not a1_address_str: # Added check for empty string        
        return a1_address_str

    try:
        row, col = coordinate_to_tuple(a1_address_str)
        target_coord = (row, col)
    except Exception as e:        
        return a1_address_str

    for merged_range in ws.merged_cells.ranges:
        # Check if the target cell (row, col) is within this merged range
        if target_coord[0] >= merged_range.min_row and target_coord[0] <= merged_range.max_row and \
           target_coord[1] >= merged_range.min_col and target_coord[1] <= merged_range.max_col:

            # Get the top-left coordinates of the merged range
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_address = f"{get_column_letter(min_col)}{min_row}"
            return top_left_address
    return a1_address_str

def inject_table1(wb_src, ws_tgt, conf, df_map, log=None):
    start_sheet = conf.get("start_sheet")
    end_sheet = conf.get("end_sheet")
    
    if not start_sheet or not end_sheet:        
        return

    try:
        ws_src_init = wb_src[start_sheet]
        ws_src_final = wb_src[end_sheet]
    except KeyError as e:        
        return

    for idx, row in df_map.iterrows():
        src_field = str(row["来源字段"]).strip()
        tgt_init_cell = str(row["目标单元格（期初）"]).strip()
        tgt_final_cell = str(row["目标单元格（期末）"]).strip()
        var_cell = str(row.get("变动单元格", "")).strip()
        var_formula = str(row.get("变动公式", "")).strip()

        # Handle empty/NaN values from mapping gracefully
        if not src_field:            
            continue
        # Only skip if ALL target cells are empty, allowing partial configurations
        if not tgt_init_cell and not tgt_final_cell and not var_cell:
            continue

        val_init, val_final = None, None
        
        # Search for initial value in ws_src_init
        found_init = False
        for r_search in range(1, ws_src_init.max_row + 1):
            name = ws_src_init.cell(row=r_search, column=1).value # Assuming source field is in column A (1)
            if name and src_field in str(name):
                val_init = ws_src_init.cell(row=r_search, column=2).value # Assuming value is in column B (2)                
                found_init = True
                break
        if not found_init:
            continue           

        # Search for final value in ws_src_final
        found_final = False
        for r_search in range(1, ws_src_final.max_row + 1):
            name = ws_src_final.cell(row=r_search, column=1).value # Assuming source field is in column A (1)
            if name and src_field in str(name):
                val_final = ws_src_final.cell(row=r_search, column=3).value # Assuming value is in column C (3)                
                found_final = True
                break
        if not found_final:
            continue            


        # --- 写入目标单元格 ---
        try:
            # Inject initial value
            if tgt_init_cell:
                actual_init_cell = _get_top_left_merged_cell_address(ws_tgt, tgt_init_cell)
                address = row.get("目标单元格（期初）") 
                if address: 
                    ws_tgt[actual_init_cell] = val_init
                    ws_tgt[actual_init_cell].number_format = '#,##0.00'          

            # Inject final value
            if tgt_final_cell:
                actual_final_cell = _get_top_left_merged_cell_address(ws_tgt, tgt_final_cell)
                address = row.get("目标单元格（期末）") 
                if address: 
                    ws_tgt[actual_final_cell] = val_final
                    ws_tgt[actual_final_cell].number_format = '#,##0.00'                

            # Inject variance formula or value
            if var_cell:
                actual_var_cell = _get_top_left_merged_cell_address(ws_tgt, var_cell)
                
                # 【核心修复】使用 if/else 分开处理
                if var_formula:
                    # 如果配置了公式，则只写入公式
                    ws_tgt[actual_var_cell] = var_formula
                    # 也可以在这里为公式单元格设置格式
                    ws_tgt[actual_var_cell].number_format = '#,##0.00'
                else:
                    # 如果没有配置公式，才进行计算并写入差值
                    num_init = float(val_init) if val_init is not None and str(val_init).strip() != "" else 0.0
                    num_final = float(val_final) if val_final is not None and str(val_final).strip() != "" else 0.0
                    
                    diff = num_final - num_init
                    ws_tgt[actual_var_cell] = diff
                    ws_tgt[actual_var_cell].number_format = '#,##0.00'                   

        except Exception as e:            
            raise