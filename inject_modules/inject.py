from openpyxl import load_workbook
from .mapping import get_mapping_conf_and_df
from .table1 import inject_table1
from .table2 import inject_table2
from .table3 import inject_table3
from .formula import inject_formula_sheet

log = []

def run_full_injection(mapping_file, source_file, template_file, output_file, log=None):
    wb_src = load_workbook(source_file, data_only=True)
    wb_tgt = load_workbook(template_file)
    ws_tgt = wb_tgt.active

    conf1, df1 = get_mapping_conf_and_df(mapping_file, "inj1")
    conf2, df2 = get_mapping_conf_and_df(mapping_file, "inj2")
    conf3, df3 = get_mapping_conf_and_df(mapping_file, "inj3")

    inject_table1(wb_src, ws_tgt, conf1, df1, log=log)
    inject_table2(wb_src, ws_tgt, conf2, df2, log=log)
    inject_table3(wb_src, ws_tgt, conf3, df3, log=log)
    inject_formula_sheet(ws_tgt, mapping_file, log=log)
   
    summary_values = {}

    wb_tgt.save(output_file)
    return summary_values