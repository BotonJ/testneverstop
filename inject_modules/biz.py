# src/biz.py

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import logging
import re

def find_correct_year_column(df_sheet: pd.DataFrame, year: str):
    """在一个业务活动表DataFrame中，根据年份标题行找到正确的金额列索引。"""
    try:
        header_row = df_sheet.iloc[3].astype(str)
        for col_idx, header_val in enumerate(header_row):
            if year in header_val:
                return col_idx
    except IndexError:
        pass
    logging.warning(f"在表头中未动态找到年份'{year}'，将默认使用第4列。")
    return 3

def get_income_expense_summary(wb_src: Workbook, mapping_file_path: str):
    """
    遍历源工作簿，根据mapping配置精确汇总收支，并生成按年份展开的透视表。
    """
    try:
        # 1. 读取科目配置
        df_raw_subjects = pd.read_excel(mapping_file_path, sheet_name="业务活动表汇总注入配置", header=None)
        header_row_index = df_raw_subjects[df_raw_subjects[0] == '类型'].index[0]
        correct_headers = df_raw_subjects.iloc[header_row_index].tolist()
        df_subjects = df_raw_subjects.iloc[header_row_index + 1:]
        df_subjects.columns = correct_headers
        income_subjects = df_subjects[df_subjects['类型'] == '收入']['科目名称'].tolist()
        expense_subjects = df_subjects[df_subjects['类型'] == '支出']['科目名称'].tolist()

        # 2. 【核心修复】读取并解析全局审计期间
        df_header = pd.read_excel(mapping_file_path, sheet_name="HeaderMapping", header=None)
        audit_period_row = df_header.loc[df_header[0] == '期末']
        audit_period_str = audit_period_row.iloc[0, 2]
        match = re.match(r'(\d{4})年(\d{1,2})月[-至](\d{4})年(\d{1,2})月', audit_period_str.replace(" ", ""))
        start_year, start_month, end_year, end_month = map(int, match.groups())
        
        logging.info(f"成功加载并解析收支配置及审计期间 ({audit_period_str})。")

    except Exception as e:
        logging.error(f"读取或解析mapping配置失败: {e}。")
        return pd.DataFrame(), pd.DataFrame(), {"收入汇总": 0.0, "支出汇总": 0.0, "收支结余汇总": 0.0}

    all_data = []
    for sheet_name in wb_src.sheetnames:
        if '业务活动表' in sheet_name:
            year_match = re.search(r'\d{4}', sheet_name)
            if not year_match: continue
            
            current_year = int(year_match.group(0))
            logging.info(f"正在处理Sheet: {sheet_name} (年份: {current_year})...")
            
            # 3. 【核心修复】应用条件格式化生成最终的项目名称
            project_name = f"{current_year}年"
            if current_year == start_year and start_month != 1:
                project_name = f"{current_year}年{start_month}-12月累计数"
            if current_year == end_year and end_month != 12:
                if start_year == end_year:
                    project_name = f"{current_year}年{start_month}-{end_month}月累计数"
                else:
                    project_name = f"{current_year}年1-{end_month}月累计数"
            
            ws = wb_src[sheet_name]
            df = pd.DataFrame(ws.values)
            if df.empty: continue

            amount_col_idx = find_correct_year_column(df, str(current_year))
            df_year_data = df.iloc[:, [0, amount_col_idx]].copy()
            df_year_data.columns = ['科目', '金额']
            df_year_data['科目'] = df_year_data['科目'].astype(str).str.replace(r'^\s*[（(].*?[)）]\s*', '', regex=True).str.strip()
            df_year_data['项目'] = project_name # <-- 使用新生成的项目名称
            all_data.append(df_year_data)

    if not all_data: return pd.DataFrame(), pd.DataFrame(), {"收入汇总": 0.0, "支出汇总": 0.0, "收支结余汇总": 0.0}

    full_df = pd.concat(all_data, ignore_index=True)
    full_df['金额'] = pd.to_numeric(full_df['金额'], errors='coerce').fillna(0)

    # --- 后续透视表逻辑使用新的“项目”列作为index ---
    def create_pivot_summary(df_filtered):
        if df_filtered.empty: return pd.DataFrame()
        pivot = pd.pivot_table(df_filtered, values='金额', index='项目', columns='科目', aggfunc='sum', fill_value=0)
        pivot['合计'] = pivot.sum(axis=1)
        pivot.loc['合计'] = pivot.sum(axis=0)
        return pivot.reset_index()

    income_summary_df = create_pivot_summary(full_df[full_df['科目'].isin(income_subjects)])
    expense_summary_df = create_pivot_summary(full_df[full_df['科目'].isin(expense_subjects)])

    total_income = income_summary_df.loc[income_summary_df['项目'] == '合计', '合计'].iloc[0] if not income_summary_df.empty else 0
    total_expense = expense_summary_df.loc[expense_summary_df['项目'] == '合计', '合计'].iloc[0] if not expense_summary_df.empty else 0

    return income_summary_df, expense_summary_df, {"收入汇总": total_income, "支出汇总": total_expense, "收支结余汇总": total_income - total_expense}


def inject_income_expense_sheets(wb_tgt: Workbook, income_df: pd.DataFrame, expense_df: pd.DataFrame):
    # 此函数保持不变
    def _populate_sheet(ws, df):
        for c_idx, value in enumerate(df.columns, 1):
            ws.cell(row=1, column=c_idx, value=value)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (int, float)): cell.number_format = '#,##0.00'

    for sheet_name, df in [("收入汇总", income_df), ("支出汇总", expense_df)]:
        if not df.empty and sheet_name in wb_tgt.sheetnames:
            ws = wb_tgt[sheet_name]
            for r in ws.iter_rows():
                for c in r: c.value = None
            _populate_sheet(ws, df)
            logging.info(f"已成功将数据注入到 '{sheet_name}' Sheet。")