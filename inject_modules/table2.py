# inject_modules/table2.py
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter

# 复制这个完整的函数
def _get_top_left_merged_cell_address(ws, a1_address_str):
    """
    Given a worksheet and an A1-style cell address string,
    returns the A1-style address of the top-left cell of the merged region it belongs to.
    If the cell is not merged, returns the original address.
    """
    if not isinstance(a1_address_str, str) or not a1_address_str:
        return a1_address_str

    try:
        row, col = coordinate_to_tuple(a1_address_str)
        target_coord = (row, col)
    except Exception as e:
        return a1_address_str

    for merged_range in ws.merged_cells.ranges:
        if target_coord[0] >= merged_range.min_row and target_coord[0] <= merged_range.max_row and \
           target_coord[1] >= merged_range.min_col and target_coord[1] <= merged_range.max_col:
            
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_address = f"{get_column_letter(min_col)}{min_row}"
            return top_left_address
    return a1_address_str

def _get_value(ws_src, row_idx, col_letter):
    """安全地获取单元格的值，如果为空则返回0。"""
    val = ws_src[f"{col_letter}{row_idx}"].value
    return float(val) if val is not None else 0.0

def inject_table2(wb_src: Workbook, ws_tgt: Worksheet, conf: dict, df_map, log=None):
    """
    根据inj2配置，注入资产和负债明细，并在每个区块下方动态生成合计行。
    """
    start_sheet_name = conf.get("start_sheet")
    end_sheet_name = conf.get("end_sheet")

    if not (start_sheet_name and end_sheet_name and start_sheet_name in wb_src.sheetnames and end_sheet_name in wb_src.sheetnames):
        
        return

    ws_start = wb_src[start_sheet_name]
    ws_end = wb_src[end_sheet_name]

    # 遍历inj2中的每个配置区块（资产区块、负债区块）
    for _, row_config in df_map.iterrows():
        
        # --- 1. 读取当前区块的配置 ---
        plate = row_config.get("区块名称", "").strip()
        start_row = int(row_config["起始行"])
        end_row = int(row_config["终止行"])
        src_col_init = str(row_config["来源列（期初）"]).strip()
        src_col_final = str(row_config["来源列（期末）"]).strip()
        
        tgt_start_cell = row_config["目标起始单元格"]
        tgt_row_cursor = int(tgt_start_cell[1:]) # 当前要写入的目标行
        tgt_col_prefix = tgt_start_cell[0].strip()

        skip_strs = [s.strip() for s in str(row_config.get("跳过行", "")).split(',') if s]
        skip_zero = str(row_config.get("是否跳过均为0", "")) == "是"
        total_name = row_config.get("合计行名称")        
        
        # 记录下第一个写入的行号，用于后续求和
        first_written_row = tgt_row_cursor

        # --- 2. 循环注入当前区块的明细行 ---
        for r_idx in range(start_row, end_row + 1):
            subject = str(ws_start.cell(row=r_idx, column=1).value).strip()

            # 如果科目需要跳过，则进入下一轮循环
            if any(s in subject for s in skip_strs if s):
                continue

            val_start = _get_value(ws_start, r_idx, src_col_init)
            val_end = _get_value(ws_end, r_idx, src_col_final)
            change = val_end - val_start

            # 如果配置了跳过0，且期初期末都为0，则跳过
            if skip_zero and val_start == 0 and val_end == 0:
                continue

            # 写入科目名称、期初、期末、变动额
            # --- 【核心修复】在每次写入前，都调用安全检查函数 ---

            # 写入科目名称
            original_addr = f"{tgt_col_prefix}{tgt_row_cursor}"
            actual_addr = _get_top_left_merged_cell_address(ws_tgt, original_addr)
            ws_tgt[actual_addr].value = subject
            
            # 写入期初值
            original_addr_start = f"{chr(ord(tgt_col_prefix)+1)}{tgt_row_cursor}"
            actual_addr_start = _get_top_left_merged_cell_address(ws_tgt, original_addr_start)
            cell_start = ws_tgt[actual_addr_start]
            cell_start.value = val_start
            cell_start.number_format = '#,##0.00'

            # 写入期末值
            original_addr_end = f"{chr(ord(tgt_col_prefix)+2)}{tgt_row_cursor}"
            actual_addr_end = _get_top_left_merged_cell_address(ws_tgt, original_addr_end)
            cell_end = ws_tgt[actual_addr_end]
            cell_end.value = val_end
            cell_end.number_format = '#,##0.00'
            
            # 写入变动额
            original_addr_change = f"{chr(ord(tgt_col_prefix)+3)}{tgt_row_cursor}"
            actual_addr_change = _get_top_left_merged_cell_address(ws_tgt, original_addr_change)
            cell_change = ws_tgt[actual_addr_change]
            cell_change.value = change
            cell_change.number_format = '#,##0.00'

            tgt_row_cursor += 1 # 目标行下移一行