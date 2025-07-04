# src/main_runner.py

import os
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# 模块导入
from modules.collector import collect_summary_values
from inject_modules.table_injector import populate_balance_change_sheet
from inject_modules.text_renderer import render_text_template_from_mapping, inject_text_to_excel
from src.legacy_runner import run_main_injection
from inject_modules.biz import get_income_expense_summary, inject_income_expense_sheets

# 粘贴在 import 之后，run_main 之前

def setup_logging(log_dir="logs", log_file="audit_autogen.log"):
    """
    配置全局日志系统，使其能同时输出到文件和终端。
    """
    # 创建logs文件夹（如果不存在）
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, log_file)

    # 1. 获取根日志记录器 (root logger)
    #    这是所有日志信息的“总源头”
    logger = logging.getLogger()
    logger.setLevel(logging.INFO) # 设置最低响应级别为INFO

    # 2. 移除所有之前可能存在的处理器，确保配置干净
    if logger.hasHandlers():
        logger.handlers.clear()

    # 3. 创建文件处理器 (FileHandler)
    #    负责将日志写入到文件中
    file_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(module)s - %(message)s')
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    # 4. 创建终端处理器 (StreamHandler)
    #    负责将日志打印到控制台
    stream_handler = logging.StreamHandler()
    stream_formatter = logging.Formatter('%(levelname)s: %(message)s') # 终端格式可以简洁一些
    stream_handler.setFormatter(stream_formatter)
    logger.addHandler(stream_handler)

    logging.info("中央日志系统已成功配置，将同时输出到文件和终端。")


def apply_global_formatting(wb, sheet_names):
    """
    遍历指定工作表，为不同类型的Sheet应用不同的、专业的财务格式。
    """
    # 为汇总表定义的格式：0显示为'-'
    summary_format = '#,##0.00;-#,##0.00;"-"'
    # 为业务活动表定义的格式：0显示为空白
    activity_format = '#,##0.00;-#,##0.00;;@'
    # 为行次列定义的整数格式
    integer_format = '0'
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    logging.info(f"开始对Sheet列表应用智能全局数字格式...")
    
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            logging.warning(f"应用全局格式化时，未找到名为'{sheet_name}'的Sheet，已跳过。")
            continue

        ws = wb[sheet_name]
        
        # --- 【核心修复】智能格式化逻辑 ---
        
        # 1. 如果是资产负债表，则完全跳过，保留其原始模板格式
        if "资产负债表" in sheet_name:
            logging.info(f"  -> 跳过'{sheet_name}'，保留原始格式。")
            continue
            
        # 2. 如果是业务活动表，应用特殊规则
        elif "业务活动表" in sheet_name:
            logging.info(f"  -> 为'{sheet_name}'应用业务活动表格式规则...")
            for row in ws.iter_rows():
                for cell in row:
                    # B列（行次列）设为整数
                    if cell.column == 2 and isinstance(cell.value, (int, float)):
                        cell.number_format = integer_format
                    # 其他数字列设为“0显示为空白”
                    elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = activity_format
                        cell.alignment = right_alignment

        # 3. 如果是其他表（即我们的汇总表），应用标准汇总格式
        else:
            logging.info(f"  -> 为'{sheet_name}'应用标准汇总格式规则...")
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = summary_format
                        cell.alignment = right_alignment


def run_main():
    setup_logging()
    # --- 1. 文件路径设置 ---
    project_root = Path(__file__).resolve().parents[1]
    mapping_path = project_root / "data" / "mapping_file.xlsx"
    source_path = project_root / "output" / "output.xlsx"
    final_path = project_root / "output" / "final_report.xlsx"

    os.makedirs(project_root / "output", exist_ok=True)
    logging.info("报表生成流程开始...")

    # --- 2. 确保"预制件"存在 ---
    if not source_path.exists():
        logging.info(f"{source_path} 未找到，首先运行 legacy_runner 生成...")
        run_main_injection()
        if not source_path.exists():
            logging.error(f"运行 legacy_runner 后仍未找到 {source_path}，终止执行。")
            return

    # --- 3. 加载"预制件" ---
    wb_final = load_workbook(source_path)
    wb_src_readonly = load_workbook(source_path, data_only=True)

    # --- 4. 核心数据收集与计算 ---
    logging.info("步骤 1: 提取原始 summary_values...")
    summary_values = collect_summary_values(mapping_path, source_path)
    
    logging.info("步骤 2: 计算原始收支汇总...")
    income_df, expense_df, biz_summary = get_income_expense_summary(wb_src_readonly, str(mapping_path))
    summary_values.update(biz_summary)
    
    # --- 5. 全局文字格式化 ---
    logging.info("步骤 3: 对所有数值进行最终格式化，用于文字注入...")
    for key in list(summary_values.keys()):
        if "增减" in key and summary_values.get(key.replace("增减", "变化方向")) == "减少":
            if isinstance(summary_values[key], (int, float)):
                summary_values[key] = abs(summary_values[key])
        
        if isinstance(summary_values[key], (int, float)):
            summary_values[key] = f"{summary_values[key]:,.2f}"

    logging.info(f"最终待注入的 summary_values (已格式化): {summary_values}")

    # --- 6. 填充工作簿 ---
    logging.info("步骤 4: 填充'资产负债变动' Sheet...")
    populate_balance_change_sheet(wb_src_readonly, wb_final, str(mapping_path))

    logging.info("步骤 5: 填充'收入汇总'和'支出汇总' Sheet...")
    inject_income_expense_sheets(wb_final, income_df, expense_df)

    logging.info("步骤 6: 渲染并注入最终说明文字...")
    rendered_text = render_text_template_from_mapping(mapping_path, summary_values, {})
    inject_text_to_excel(wb_final, sheet_name="支出汇总", cell="H1", text=rendered_text)

    # --- 7. 应用全局表格格式化 ---
    sheets_to_format = ["资产负债变动", "收入汇总", "支出汇总"]
    for sheet in wb_final:
        if "资产负债表" in sheet.title or "业务活动表" in sheet.title:
            if sheet.title not in sheets_to_format:
                sheets_to_format.append(sheet.title)
                
    apply_global_formatting(wb_final, sheets_to_format)
    
    # --- 8. 另存为最终报告 ---
    try:
        wb_final.save(final_path)
        logging.info(f"✅ 报表已完成，所有内容已写入：{final_path}")
    except Exception as e:
        logging.error(f"保存最终报告 {final_path} 时出错: {e}")

if __name__ == '__main__':
    run_main()