# modules/collector.py

import json
import logging
from openpyxl import load_workbook
from modules.mapping_loader import load_mapping_file
from inject_modules.balance_utils import get_balance_core_data
import re
import calendar

def get_change_direction(summary: dict):
    # 此函数保持不变
    fields_to_check = {
        "资产总额": "资产变化方向", "负债总额": "负债变化方向", "净资产总额": "净资产变化方向"
    }
    for base_field, direction_field in fields_to_check.items():
        change_key = f"{base_field}增减"
        change_value = summary.get(change_key)
        try:
            val_to_compare = 0.0
            if isinstance(change_value, (int, float)):
                val_to_compare = float(change_value)
            elif isinstance(change_value, str):
                if "计算失败" in change_value:
                    summary[direction_field] = "【无法计算】"
                    continue
                val_to_compare = float(change_value.replace(",", "").strip())
            
            if val_to_compare > 0: summary[direction_field] = "增长"
            elif val_to_compare < 0: 
                summary[direction_field] = "减少"
                summary[change_key] = abs(val_to_compare) # 修复“减少-负数”问题
            else: summary[direction_field] = "保持不变"
        except (ValueError, TypeError):
            summary[direction_field] = "【无法计算】"

def collect_summary_values(mapping_path, output_path):
    summary = {}
    # ... (前面的 mapping 和 alias_dict 加载逻辑保持不变) ...
    mapping = load_mapping_file(mapping_path)
    raw_alias_map = mapping["subject_alias_map"]
    alias_dict = {}
    for std, aliases in raw_alias_map.items():
        std_norm = std.strip()
        if not isinstance(aliases, list): aliases = [aliases]
        for alias in [std_norm] + aliases:
            alias_norm = alias.strip()
            alias_dict[alias_norm] = std_norm
            
    try:
        mapping_wb = load_workbook(mapping_path, data_only=True)
        header_ws = mapping_wb["HeaderMapping"]
        rule_dict = {
            row[0].value: str(row[2].value).strip() if row[2].value is not None else ""
            for row in header_ws.iter_rows(min_row=2)
        }

        # 提取数据，但不进行格式化
        unit_name = rule_dict.get("单位名称", "【未提取】")
        summary["单位名称"] = unit_name.replace("编制单位：", "").strip()
        summary["审计期间"] = rule_dict.get("期末", "【未提取】")
        
        audit_period_str = summary.get('审计期间')
        if audit_period_str:
            match = re.match(r'(\d{4})年(\d{1,2})月[-至](\d{4})年(\d{1,2})月', audit_period_str.replace(" ", ""))
            if match:
                start_year, start_month, end_year, end_month = map(int, match.groups())
                summary['起始日期'] = f"{start_year}年{start_month}月1日"
                _, last_day = calendar.monthrange(end_year, end_month)
                summary['终止日期'] = f"{end_year}年{end_month}月{last_day}日"
        
        start_sheet = rule_dict.get("起始资产负债表Sheet")
        end_sheet = rule_dict.get("终止资产负债表Sheet")
        wb = load_workbook(output_path, data_only=True)
        if start_sheet in wb.sheetnames and end_sheet in wb.sheetnames:
            start_data = get_balance_core_data(wb[start_sheet], mapping["blocks"], alias_dict)
            end_data = get_balance_core_data(wb[end_sheet], mapping["blocks"], alias_dict)
            for field in ["资产总额", "负债总额", "净资产总额"]:
                start_val = float(start_data.get(f"期初{field}", 0) or 0)
                end_val = float(end_data.get(f"期末{field}", 0) or 0)
                summary[f"期初{field}"] = start_val
                summary[f"期末{field}"] = end_val
                summary[f"{field}增减"] = round(end_val - start_val, 2)
            get_change_direction(summary)

    except Exception as e:
        logging.error(f"在 collect_summary_values 中发生错误: {e}")

    # 函数结束，返回的是包含原始数字的字典
    return summary