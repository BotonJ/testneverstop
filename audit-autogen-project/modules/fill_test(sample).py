# FillTemplateByMapping.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import os
from copy import copy

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# 文件路径
mapping_file = Path("mapping_file.xlsx")
source_file = Path("soce.xlsx")
target_file = Path("t.xlsx")
output_file = Path("输出报表.xlsx")
log_file = Path("填报日志.txt")

# 加载映射表
block_map_df = pd.read_excel(mapping_file, sheet_name="资产负债表区块")
biz_map_df = pd.read_excel(mapping_file, sheet_name="业务活动表逐行")
header_map_df = pd.read_excel(mapping_file, sheet_name=2)  # 表头与公司名称映射
try:
    alias_df = pd.read_excel(mapping_file, sheet_name="科目等价映射")
    alias_dict = {}
    for _, row in alias_df.iterrows():
        std = str(row["标准科目名"]).strip()
        for col in row.index:
            if col == "标准科目名":
                continue
            alt = str(row[col]).strip()
            if alt and alt.lower() != "nan":
                alias_dict[alt] = std
except Exception:
    alias_dict = {}

# 加载工作簿
wb_source = load_workbook(source_file, data_only=True)
wb_template = load_workbook(target_file)
wb_output = wb_template  # 保证 copy_worksheet 在同一工作簿内

log = []

def get_year_from_sheet(sheet_name):
    m = re.match(r"(\d{4})", sheet_name)
    return int(m.group(1)) if m else None

def map_alias(name):
    name = str(name).strip()
    return alias_dict.get(name, name)

def write_to_cells(ws, cells_str, value):
    for cell in str(cells_str).split(","):
        cell = cell.strip()
        if cell:
            ws[cell] = value
            ws[cell].number_format = "@"  # 强制文本，避免日期格式错乱

def fill_headers_and_names(wb_output, sheet_name, year):
    for _, row in header_map_df.iterrows():
        fill_type = str(row["类型"]).strip()
        name = str(row["名称"]).strip()
        tgt_cells_b = str(row["目标资产负债表单元格"]).strip()
        tgt_cells_y = str(row["目标业务活动表单元格"]).strip()
        if fill_type == "公司名称":
            if tgt_cells_b and sheet_name.endswith("资产负债表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_b, name)
            if tgt_cells_y and sheet_name.endswith("业务活动表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_y, name)
        elif fill_type == "期初表头":
            val = f"{year - 1}年12月31日" if year else ""
            if tgt_cells_b and sheet_name.endswith("资产负债表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_b, val)
            if tgt_cells_y and sheet_name.endswith("业务活动表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_y, val)
        elif fill_type == "期末表头":
            val = f"{year}年12月31日" if year else ""
            if tgt_cells_b and sheet_name.endswith("资产负债表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_b, val)
            if tgt_cells_y and sheet_name.endswith("业务活动表"):
                write_to_cells(wb_output[sheet_name], tgt_cells_y, val)

def fill_balance_sheet(ws_src, ws_tgt, block_map_df):
    for _, block in block_map_df.iterrows():
        start_cell = str(block["起始单元格"]).strip()
        end_cell = str(block["终止单元格"]).strip()
        tgt_start_cell = str(block.get("目标起始单元格", "")).strip()
        tgt_end_cell = str(block.get("目标终止单元格", "")).strip()
        src_col_init = str(block.get("源期初列", "")).strip()
        src_col_final = str(block.get("源期末列", "")).strip()
        tgt_col_init = str(block.get("目标期初列", "")).strip()
        tgt_col_final = str(block.get("目标期末列", "")).strip()
        skip_strs = str(block.get("跳过行", "")).split(",")

        if not (start_cell and end_cell and tgt_start_cell and tgt_end_cell):
            log.append(f"⚠️ 区块定义不完整，跳过：{start_cell}-{end_cell}")
            continue

        src_start = int(re.findall(r"\d+", start_cell)[0])
        src_end = int(re.findall(r"\d+", end_cell)[0])
        tgt_start = int(re.findall(r"\d+", tgt_start_cell)[0])
        tgt_end = int(re.findall(r"\d+", tgt_end_cell)[0])

        row_count = min(src_end - src_start + 1, tgt_end - tgt_start + 1)
        log.append(f"🔄 区块对齐：{start_cell}-{end_cell} → {tgt_start_cell}-{tgt_end_cell}，共 {row_count} 行")

        for offset in range(row_count):
            src_i = src_start + offset
            tgt_i = tgt_start + offset
            name = ws_src[f"A{src_i}"].value
            if not isinstance(name, str) or name.strip() == "":
                continue
            name_norm = map_alias(name)
            if any(s.strip() in name_norm for s in skip_strs):
                continue
            try:
                val_init = ws_src[f"{src_col_init}{src_i}"].value if src_col_init else ""
                val_final = ws_src[f"{src_col_final}{src_i}"].value if src_col_final else ""
                ws_tgt[f"{tgt_col_init}{tgt_i}"] = val_init if val_init is not None else ""
                ws_tgt[f"{tgt_col_final}{tgt_i}"] = val_final if val_final is not None else ""
                log.append(f"✅ {name_norm} → {tgt_col_init}{tgt_i}/{tgt_col_final}{tgt_i}：{val_init}/{val_final}")
            except Exception as e:
                log.append(f"⚠️ {name_norm} 第 {src_i} 行写入失败：{e}")

def fill_biz_sheet(ws_src, ws_tgt, ws_tgt_name, all_ws_biz, biz_map_df):
    year = get_year_from_sheet(ws_tgt_name)
    prev_ws_tgt = next((ws for ws in all_ws_biz if get_year_from_sheet(ws.title) == year - 1), None)

    for _, row in biz_map_df.iterrows():
        subject = map_alias(row["科目名称"])
        src_final_cell = str(row.get("源期末坐标", "")).strip()
        tgt_init_cell = str(row.get("目标期初坐标", "")).strip()
        tgt_final_cell = str(row.get("目标期末坐标", "")).strip()
        
        if tgt_init_cell and prev_ws_tgt and tgt_final_cell:
            try:
                prev_val = prev_ws_tgt[tgt_final_cell].value
                ws_tgt[tgt_init_cell] = prev_val if prev_val is not None else ""
                log.append(f"✅ [{year}] {subject} 期初 ← 上年{year-1}期末 {tgt_final_cell} → {tgt_init_cell}：{prev_val}")
            except Exception as e:
                log.append(f"⚠️ [{year}] {subject} 期初写入失败：{e}")

        if tgt_final_cell:
            try:
                val = ""
                cell = src_final_cell.strip()
                if ws_src:
                    try:
                        val = ws_src[cell].value
                        log.append(f"🧪 [{year}] {subject} 直接读取 {cell} → {val}")
                    except Exception as e:
                        log.append(f"⚠️ [{year}] {subject} 源坐标读取失败 {cell}：{e}")
                else:
                    log.append(f"⚠️ [{year}] {subject} 源工作表未定义或丢失：{cell}")
                ws_tgt[tgt_final_cell] = val if val is not None else ""
                log.append(f"✅ [{year}] {subject} 期末 → {tgt_final_cell}：{val}")
            except Exception as e:
                log.append(f"⚠️ [{year}] {subject} 期末写入失败 {tgt_final_cell}，错误：{e}")

# --- 生成所有年份sheet ---
years = sorted({get_year_from_sheet(name) for name in wb_source.sheetnames if get_year_from_sheet(name)})
tmpl_sheets = ["资产负债表", "业务活动表"]
for wsname in list(wb_output.sheetnames):
    if wsname not in tmpl_sheets:
        wb_output.remove(wb_output[wsname])

for year in years:
    for base in tmpl_sheets:
        title = f"{year}{base}"
        if title not in wb_output.sheetnames:
            tmpl_sheet = wb_output[base]
            new_sheet = wb_output.copy_worksheet(tmpl_sheet)
            new_sheet.title = title
for tmpl in tmpl_sheets:
    if tmpl in wb_output.sheetnames:
        wb_output.remove(wb_output[tmpl])

# === 主流程 ===
all_ws_biz = [ws for ws in wb_output.worksheets if ws.title.endswith("业务活动表")]
for sheet_name in wb_output.sheetnames:
    year = get_year_from_sheet(sheet_name)
    fill_headers_and_names(wb_output, sheet_name, year)
    if sheet_name.endswith("资产负债表"):
        src_sheet_name = f"{year}资产负债表"
        if src_sheet_name not in wb_source.sheetnames:
            log.append(f"⚠️ 源文件缺失：{src_sheet_name}")
            continue
        fill_balance_sheet(wb_source[src_sheet_name], wb_output[sheet_name], block_map_df)
    elif sheet_name.endswith("业务活动表"):
        if sheet_name not in wb_source.sheetnames:
            log.append(f"⚠️ 源文件缺失：{sheet_name}")
            continue
        fill_biz_sheet(wb_source[sheet_name], wb_output[sheet_name], sheet_name, all_ws_biz, biz_map_df)

wb_output.save(output_file)
with open(log_file, "w", encoding="utf-8") as f:
    f.write("\n".join(log))
print(f"✅ 已完成填报：{output_file}")
print(f"📝 日志保存于：{log_file}")
