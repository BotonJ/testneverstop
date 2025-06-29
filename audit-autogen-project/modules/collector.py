# collector.py：统一提取 summary_values
import json
import logging
from openpyxl import load_workbook
from modules.mapping_loader import load_mapping_file
from inject_modules.balance_utils import get_balance_core_data

# Configure logging for debugging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# --- 新增函数：get_change_direction ---
def get_change_direction(summary: dict):
    """
    根据资产、负债、净资产的增减额，计算并设置其变化方向（增长/减少/保持不变）。
    结果直接更新到传入的 summary 字典中。
    """
    fields_to_check = {
        "资产总额": "资产变化方向",
        "负债总额": "负债变化方向",
        "净资产总额": "净资产变化方向"
    }

    for base_field, direction_field in fields_to_check.items():
        change_key = f"{base_field}增减"
        change_value = summary.get(change_key)
        
        try:
            val_to_compare = 0.0
            if isinstance(change_value, (int, float)):
                val_to_compare = float(change_value)
            elif isinstance(change_value, str):
                # 尝试清理字符串进行转换，同时处理可能的“计算失败”情况
                if "计算失败" in change_value:
                    summary[direction_field] = "【无法计算】"
                    #logging.warning(f"无法计算 '{direction_field}'，'{change_key}' 值: '{change_value}' 为计算失败。")
                    continue # 跳过当前字段，处理下一个
                val_to_compare = float(change_value.replace(",", "").strip())
            else:
                val_to_compare = 0.0 # 如果是 None 或其他非预期类型，默认为 0

            if val_to_compare > 0:
                summary[direction_field] = "增长"
            elif val_to_compare < 0:
                summary[direction_field] = "减少"
            else:
                summary[direction_field] = "保持不变"
                
        except (ValueError, TypeError):
            # 捕获转换失败的情况，例如字符串无法解析为数字
            summary[direction_field] = "【无法计算】"
            #logging.warning(f"无法计算 '{direction_field}'，'{change_key}' 值: '{change_value}' 无法转换为数字。")

# --- get_change_direction 函数定义结束 ---


def collect_summary_values(mapping_path, output_path):
    """
    从 mapping_file 的 HeaderMapping 中提取单位名、起止sheet；
    从 output.xlsx 中读取资产/负债/净资产期初期末值，并计算差额；
    返回完整 summary_values 字典
    """
    summary = {}
    mapping = load_mapping_file(mapping_path)
    # ✅ 插入 alias_dict 构造逻辑（反向映射）
    raw_alias_map = mapping["subject_alias_map"]
    alias_dict = {}
    for std, aliases in raw_alias_map.items():
        std_norm = std.strip()
        # ✅ 统一转为列表（无论 aliases 是 str 还是 list）
        if not isinstance(aliases, list):
            aliases = [aliases]
        for alias in [std_norm] + aliases:
            alias_norm = alias.strip()
            alias_dict[alias_norm] = std_norm
    try:
        mapping_wb = load_workbook(mapping_path, data_only=True)
        header_ws = mapping_wb["HeaderMapping"]

        # 获取 HeaderMapping 规则字段
        rule_dict = {
            row[0].value: str(row[2].value).strip() if row[2].value is not None else ""
            for row in header_ws.iter_rows(min_row=2)
        }

        summary["单位名称"] = rule_dict.get("单位名称", "【未提取】")
        summary["审计期间"] = rule_dict.get("期末", "【未提取】")

        start_sheet = rule_dict.get("起始资产负债表Sheet")
        end_sheet = rule_dict.get("终止资产负债表Sheet")

        # Log extracted sheet names for debugging
        #logging.info(f"HeaderMapping - 起始资产负债表Sheet: {start_sheet}")
        #logging.info(f"HeaderMapping - 终止资产负债表Sheet: {end_sheet}")

        # 从 output.xlsx 提取资产负债字段
        wb = load_workbook(output_path, data_only=True)
        if start_sheet in wb.sheetnames and end_sheet in wb.sheetnames:
            block_map = mapping["blocks"]
            #print("🧾 当前 block_map 内容如下：")
            #print(json.dumps(block_map, ensure_ascii=False, indent=2))                
            start_data = get_balance_core_data(wb[start_sheet], block_map, alias_dict)
            end_data = get_balance_core_data(wb[end_sheet], block_map, alias_dict)

            # Log core data for debugging
            logging.info(f"Start Sheet Data ({start_sheet}): {start_data}")
            #logging.info(f"End Sheet Data ({end_sheet}): {end_data}")

            for field in ["资产总额", "负债总额", "净资产总额"]:
                start_val = float(start_data.get(f"期初{field}", 0) or 0)
                end_val = float(end_data.get(f"期末{field}", 0) or 0)

                summary[f"期初{field}"] = start_val
                summary[f"期末{field}"] = end_val
                try:
                    summary[f"{field}增减"] = round(end_val - start_val, 2)
                except Exception as e:
                    summary[f"{field}增减"] = f"计算失败: {e}"
            
            # --- 函数调用：在所有“增减”字段计算完成后，调用 get_change_direction ---
            get_change_direction(summary)
            # --- 函数调用结束 ---

        else:
            summary["期初资产总额"] = summary["期末资产总额"] = "【未找到起止Sheet】"
            summary["期初负债总额"] = summary["期末负债总额"] = "【未找到起止Sheet】"
            summary["期初净资产总额"] = summary["期末净资产总额"] = "【未找到起止Sheet】"
            summary["资产总额增减"] = summary["负债总额增减"] = summary["净资产总额增减"] = "【未找到起止Sheet】"
            # ✅ 确保这里也添加了“变化方向”的默认值，以防模板报错
            summary["资产变化方向"] = "【未找到起止Sheet】"
            summary["负债变化方向"] = "【未找到起止Sheet】"
            summary["净资产变化方向"] = "【未找到起止Sheet】"
            #logging.warning(f"Required sheets not found in {output_path}. Start: '{start_sheet}', End: '{end_sheet}'.")

    except Exception as e:
        logging.error(f"Error in collect_summary_values: {e}")
        # Populate summary with error indicators in case of a global failure
        for key in ["单位名称", "审计期间", "期初资产总额", "期末资产总额", "资产总额增减",
                     "期初负债总额", "期末负债总额", "负债总额增减",
                     "期初净资产总额", "期末净资产总额", "净资产总额增减",
                     "资产变化方向", "负债变化方向", "净资产变化方向"]: # <-- 添加所有变化方向字段到错误处理中
            if key not in summary:
                summary[key] = f"【提取失败: {e}】"
    return summary