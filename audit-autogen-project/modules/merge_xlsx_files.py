import os
import copy
from openpyxl import load_workbook, Workbook

# 设置当前目录
os.chdir(os.path.dirname(os.path.abspath(__file__)))

output_file = "merged.xlsx"
wb_out = Workbook()
wb_out.remove(wb_out.active)

def clone_sheet(src_ws, tgt_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            tgt_cell = tgt_ws[cell.coordinate]
            tgt_cell.value = cell.value
            if cell.has_style:
                tgt_cell.font = copy.copy(cell.font)
                tgt_cell.border = copy.copy(cell.border)
                tgt_cell.fill = copy.copy(cell.fill)
                tgt_cell.number_format = copy.copy(cell.number_format)
                tgt_cell.protection = copy.copy(cell.protection)
                tgt_cell.alignment = copy.copy(cell.alignment)
    for merged_range in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged_range))

for fname in os.listdir():
    if fname.lower().endswith(".xlsx") and not fname.startswith("~$") and fname != output_file:
        wb_src = load_workbook(fname, data_only=True)
        for sheet_name in wb_src.sheetnames:
            src_ws = wb_src[sheet_name]
            target_sheet_name = sheet_name
            i = 1
            while target_sheet_name in wb_out.sheetnames:
                target_sheet_name = f"{sheet_name}_{i}"
                i += 1
            tgt_ws = wb_out.create_sheet(title=target_sheet_name)
            clone_sheet(src_ws, tgt_ws)
            print(f"✅ 已合并：{fname} → {target_sheet_name}")

wb_out.save(output_file)
print(f"\n🎉 所有文件合并完毕，输出：{output_file}")
