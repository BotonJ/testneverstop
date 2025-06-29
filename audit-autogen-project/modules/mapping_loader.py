import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple, coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter 

def get_col_index(cell):
    try:
        if cell is None:
            return None
        if cell.isalpha():  # 直接是 'C' 这种
            return column_index_from_string(cell)
        # 是坐标，如 'C7' → 取出 C → 转为列号 3
        return column_index_from_string(coordinate_from_string(cell)[0])
    except Exception as e:
        print(f"列坐标解析失败: {cell} → {e}")
        return None

def parse_skip_rows(value):
    if not value:
        return []
    rows = []
    for item in str(value).split(","):
        item = item.strip().replace("：", "")  # 删除全角冒号
        if item.isdigit():
            rows.append(int(item))
    return rows

def load_mapping_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # blocks 区块
    block_sheet = wb["资产负债表区块"]
    blocks = {}
    for row in block_sheet.iter_rows(min_row=2, values_only=True):
        block_name, start, end, sci, scf, tci, tcf, skip = row[:8]

        if not block_name or not str(block_name).strip():
            continue  # 跳过空行或无效区块名
        block_name = str(block_name).strip()

        target_start_cell_str = row[7] if len(row) > 7 else None #读取mapping文件的第8列用于汇总部分填写资产总额、负债总额等合计数
        target_row = None
        if target_start_cell_str:
            try:
                _, target_row = coordinate_from_string(str(target_start_cell_str))
            except Exception as e:
                print(f"⚠️ 目标单元格 '{target_start_cell_str}' 无法解析为 target_row: {e}")
       
        blocks[block_name] = {
            "start_row": coordinate_to_tuple(start)[0] if start else None,
            "end_row": coordinate_to_tuple(end)[0] if end else None,
            "sorce_col_initial": get_col_index(sci),
            "sorce_col_final": get_col_index(scf),
            "target_col_initial": get_col_index(tci),
            "target_col_final": get_col_index(tcf),
            "target_row": target_row,
            "skip_rows": parse_skip_rows(skip)
        }        
    # subject alias
    alias_sheet = wb["科目等价映射"]
    alias_map = {}
    for row in alias_sheet.iter_rows(min_row=2, values_only=True):
        key, *aliases = row
        alias_map[key.strip()] = [a.strip() for a in aliases if a]

    # yewu line map
    yewu_sheet = wb["业务活动表逐行"]
    yewu_map = []
    headers = [cell.value for cell in next(yewu_sheet.iter_rows(min_row=1, max_row=1))]
    for row in yewu_sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # ✅ 跳过空行
        item = dict(zip(headers, row))
        yewu_map.append(item)

    # header
    header_sheet = wb["HeaderMapping"]
    
# 读取 HeaderMapping 表头配置
    header_meta = {}
    if "HeaderMapping" in wb.sheetnames:
        ws_header = wb["HeaderMapping"]
        for row in ws_header.iter_rows(min_row=2, values_only=True):
            if not row or not str(row[0]).strip():
                continue  # ⛔ 跳过空行或列数不足的行
            name, typ, rule, balance_cells, activity_cells = row[:5]
             # ✅ 跳过空字段名，避免误把规则当成坐标处理
            if not name or str(name).strip() == "":
                continue

            entry = {
                "type": str(typ or "").strip(),
                "rule": str(rule or "").strip(),
                "target_cells": {
                    "资产负债表": [],
                    "业务活动表": []
                }
            }

            for cell in str(balance_cells or "").split(","):  
                cell = cell.strip()
                if not cell:
                    continue
                try:
                    if cell.isalpha():  # 如 "B"
                        coord = (1, column_index_from_string(cell))  # 默认第1行
                    else:
                        coord = coordinate_to_tuple(cell)
                        entry["target_cells"]["资产负债表"].append(coord)
                except Exception as e:
                    print(f"非法单元格坐标: '{cell}' 被跳过 → {e}")

            for cell in str(activity_cells or "").split(","):
                cell = cell.strip()
                if not cell:
                    continue
                try:
                    if cell.isalpha():  # 如 "B"
                        coord = (1, column_index_from_string(cell))  # 默认第1行
                    else:
                        coord = coordinate_to_tuple(cell)
                    entry["target_cells"]["业务活动表"].append(coord)
                except Exception as e:
                    print(f"非法单元格坐标: '{cell}' 被跳过 → {e}")
            header_meta[str(name).strip()] = entry            
    # ✅ HeaderMapping 逻辑结束
    # ✅ 返回结果写在函数内部
    return {
        "blocks": blocks,
        "subject_alias_map": alias_map,
        "yewu_line_map": yewu_map,
        "header_meta": header_meta
    }
