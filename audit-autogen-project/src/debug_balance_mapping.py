from openpyxl import load_workbook
from pathlib import Path
from collections import OrderedDict
from modules.mapping_loader import load_mapping_file
from modules.utils import normalize_name

# 路径配置
data_dir = Path(r"D:\python脚本合集\0审计自动化——审计报告生成\测试全过程\3.6-11资产负债表注入测试\data")  # 修改为本地路径
wb_src = load_workbook(data_dir / "soce.xlsx", data_only=True)
wb_tgt = load_workbook(data_dir / "t.xlsx")
mapping = load_mapping_file(data_dir / "mapping_file.xlsx")
alias_dict = {a: k for k, v in mapping["subject_alias_map"].items() for a in [k] + v}

def extract_combined_subjects(ws, alias_dict):
    result = OrderedDict()
    for i in range(1, ws.max_row + 1):
        name_a = ws[f"A{i}"].value
        if name_a:
            name_std = normalize_name(alias_dict.get(str(name_a).strip(), str(name_a).strip()))
            val_init = ws[f"C{i}"].value or ""
            val_final = ws[f"D{i}"].value or ""
            result[f"{name_std} (A列)"] = {"期初": val_init, "期末": val_final, "行号": i}
        name_e = ws[f"E{i}"].value
        if name_e:
            name_std = normalize_name(alias_dict.get(str(name_e).strip(), str(name_e).strip()))
            val_init = ws[f"G{i}"].value or ""
            val_final = ws[f"H{i}"].value or ""
            result[f"{name_std} (E列)"] = {"期初": val_init, "期末": val_final, "行号": i}
    return result

def extract_template_subjects(ws):
    result = {}
    for i in range(1, ws.max_row + 1):
        name = ws[f"A{i}"].value
        if name:
            std_name = normalize_name(str(name).strip())
            result[std_name] = i
    return result

# 主流程
for sheet_name in wb_src.sheetnames:
    if "资产负债表" not in sheet_name:
        continue
    year = int(sheet_name[:4])
    ws_src = wb_src[sheet_name]
    ws_tgt = wb_tgt.copy_worksheet(wb_tgt["资产负债表"])
    ws_tgt.title = f"{year}资产负债表"

    src_dict = extract_combined_subjects(ws_src, alias_dict)
    tgt_dict = extract_template_subjects(ws_tgt)

    print(f"\n🧾 年度：{year}")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    for name, info in src_dict.items():
        base_name = name.replace(" (A列)", "").replace(" (E列)", "")
        match = [k for k in tgt_dict.keys() if base_name == k]
        match_row = tgt_dict[match[0]] if match else ""
        print(f"📌 科目: {name} | 行号: {info['行号']} | 期初: {info['期初']} | 期末: {info['期末']} | 模板行: {match_row}")
