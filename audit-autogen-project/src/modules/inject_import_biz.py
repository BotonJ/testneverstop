import os
from openpyxl import load_workbook
import pandas as pd
from inject_biz_summary import inject_summary_fuzzy
from openpyxl.utils.cell import get_column_letter
from inject_biz_summary import inject_summary_fuzzy

os.chdir(os.path.dirname(os.path.abspath(__file__)))
# 文件路径（可根据需要手动修改为绝对路径）
mapping_file = "mapping_file.xlsx"
source_file = "source_file.xlsx"
template_file = "template_file.xlsx"
output_file = "output_file.xlsx"

# 读取字段配置
raw_df = pd.read_excel(mapping_file, sheet_name="业务活动表汇总注入配置", header=None)
conf = {}
data_start = 0
for i, row in raw_df.iterrows():
    key = str(row[0]).strip() if pd.notna(row[0]) else ""
    val = str(row[1]).strip() if pd.notna(row[1]) else ""
    if key in ["start_sheet", "end_sheet"]:
        conf[key] = val
    if any(str(cell).strip() in ["类型", "科目名称"] for cell in row if pd.notna(cell)):
        data_start = i
        break

# 提取收入和支出科目
df_mapping = pd.read_excel(mapping_file, sheet_name="业务活动表汇总注入配置", header=data_start)
income_subjects = df_mapping[df_mapping["类型"] == "收入"]["科目名称"].astype(str).tolist()
expense_subjects = df_mapping[df_mapping["类型"] == "支出"]["科目名称"].astype(str).tolist()

# 打开源工作簿和模板
wb_src = load_workbook(source_file, data_only=True)
wb_tgt = load_workbook(template_file)

# 提取 sheet 范围
sheets = wb_src.sheetnames
s_idx = sheets.index(conf["start_sheet"])
e_idx = sheets.index(conf["end_sheet"])
sheets_to_process = [s for s in sheets[s_idx:e_idx+1] if "资产负债表" not in s]

# 写入收入汇总
ws_income = wb_tgt.create_sheet("收入汇总")
inject_summary_fuzzy(income_subjects, ws_income, sheets_to_process, wb_src, conf)

# 写入支出汇总
ws_expense = wb_tgt.create_sheet("支出汇总")
inject_summary_fuzzy(expense_subjects, ws_expense, sheets_to_process, wb_src, conf)

# 保存输出
wb_tgt.save(output_file)
print("写入完成，输出文件已保存为：", output_file)
