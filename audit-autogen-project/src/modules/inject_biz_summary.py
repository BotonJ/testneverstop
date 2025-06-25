import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
os.chdir(os.path.dirname(os.path.abspath(__file__)))

def inject_summary_fuzzy(subjects_whitelist, ws_tgt, sheets_to_process, wb_src, conf):
    row_offset = 2
    col_offset = 2
    subject_col_map = {}
    col_cursor = col_offset

    ws_tgt.cell(row=1, column=1).value = "项目"

    for i, sheet in enumerate(sheets_to_process):
        ws_src = wb_src[sheet]
        d_header = ws_src.cell(row=4, column=4).value
        year_str = str(d_header).strip() if sheet == conf["end_sheet"] else str(d_header).replace("累计数", "").strip()
        row_tgt = row_offset + i
        ws_tgt.cell(row=row_tgt, column=1).value = year_str

        for r in range(5, ws_src.max_row + 1):
            raw_subject = str(ws_src.cell(row=r, column=1).value).strip()
            value = ws_src.cell(row=r, column=4).value

            matched_name = next((std for std in subjects_whitelist if std in raw_subject), None)
            if matched_name and value not in [None, "", 0, 0.0]:
                if matched_name not in subject_col_map:
                    subject_col_map[matched_name] = col_cursor
                    ws_tgt.cell(row=1, column=col_cursor).value = matched_name
                    col_cursor += 1

                col = subject_col_map[matched_name]
                ws_tgt.cell(row=row_tgt, column=col).value = float(value)
                ws_tgt.cell(row=row_tgt, column=col).number_format = '#,##0.00'

    last_col = max(subject_col_map.values(), default=col_offset)

    for r in range(row_offset, row_offset + len(sheets_to_process)):
        formula = f"=SUM({get_column_letter(col_offset)}{r}:{get_column_letter(last_col)}{r})"
        ws_tgt.cell(row=r, column=last_col + 1).value = formula
        ws_tgt.cell(row=r, column=last_col + 1).number_format = '#,##0.00'
    ws_tgt.cell(row=1, column=last_col + 1).value = "合计"

    total_row = row_offset + len(sheets_to_process)
    ws_tgt.cell(row=total_row, column=1).value = "合计"
    for col in range(col_offset, last_col + 2):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{row_offset}:{col_letter}{total_row - 1})"
        ws_tgt.cell(row=total_row, column=col).value = formula
        ws_tgt.cell(row=total_row, column=col).number_format = '#,##0.00'

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='inject_biz_summary.log',
    filemode='w'
)

def apply_number_format(cell):
    cell.number_format = '#,##0.00'

def clean_numeric(value):
    if isinstance(value, str):
        value = value.replace(',', '').replace('¥', '').strip()
    try:
        return float(value) if value not in [None, ''] else 0.0
    except:
        return 0.0

def inject_biz_summary(
    mapping_file,
    source_file,
    template_file,
    output_file,
    ws_tgt_name='Sheet1',
    config_sheet='业务活动表汇总注入配置'
):
    try:
        # 读取配置
        raw_df = pd.read_excel(mapping_file, sheet_name=config_sheet, header=None)
        conf = {}
        data_start = 0
        for i, row in raw_df.iterrows():
            key = str(row[0]).strip() if pd.notna(row[0]) else ""
            val = str(row[1]).strip() if pd.notna(row[1]) else ""
            if key in ["start_sheet", "end_sheet"]:
                conf[key] = val
            if any(str(cell).strip() in ["类型", "科目名称"] for cell in row if pd.notna(cell)):
                data_start = i
                break

        df_mapping = pd.read_excel(mapping_file, sheet_name=config_sheet, header=data_start)
        allowed_subjects = set(df_mapping['科目名称'].dropna().astype(str))

        start_sheet = conf.get("start_sheet")
        end_sheet = conf.get("end_sheet")

        wb_src = load_workbook(source_file, data_only=True)
        wb_tgt = load_workbook(template_file)
        ws_tgt = wb_tgt.create_sheet("业务活动汇总", index=len(wb_tgt.sheetnames))

        # 第一行第一列写入“项目”标题
        ws_tgt.cell(row=1, column=1).value = "项目"

        sheets = wb_src.sheetnames
        s_idx = sheets.index(start_sheet)
        e_idx = sheets.index(end_sheet)
        sheets_to_process = [s for s in sheets[s_idx:e_idx+1] if "资产负债表" not in s]

        row_offset = 2
        col_offset = 2
        subject_col_map = {}
        col_cursor = col_offset

        for i, sheet_name in enumerate(sheets_to_process):
            ws_src = wb_src[sheet_name]
            max_row = ws_src.max_row

            d_header = ws_src.cell(row=4, column=4).value
            # 处理末尾 sheet 名显示特殊规则
            if sheet_name == end_sheet:
                year_name = str(d_header) if d_header else sheet_name.replace("业务活动表", "")
            else:
                year_name = str(d_header).replace("累计数", "") if d_header else sheet_name.replace("业务活动表", "")
            row_tgt = row_offset + i
            ws_tgt.cell(row=row_tgt, column=1).value = year_name

            for r in range(5, max_row + 1):
                subject = ws_src.cell(row=r, column=1).value
                value = clean_numeric(ws_src.cell(row=r, column=4).value)
                if not subject or value == 0.0:
                    continue
                if str(subject).strip() not in allowed_subjects:
                    continue

                if subject not in subject_col_map:
                    subject_col_map[subject] = col_cursor
                    ws_tgt.cell(row=1, column=col_cursor).value = subject
                    col_cursor += 1

                col_idx = subject_col_map[subject]
                ws_tgt.cell(row=row_tgt, column=col_idx).value = value
                apply_number_format(ws_tgt.cell(row=row_tgt, column=col_idx))

        if not subject_col_map:
            wb_tgt.save(output_file)
            logging.warning("无有效数据写入，输出为空")
            return

        last_col = max(subject_col_map.values())

        # 设置合计标题列
        ws_tgt.cell(row=1, column=last_col + 1).value = "合计"

        for r in range(row_offset, row_offset + len(sheets_to_process)):
            total_cell = ws_tgt.cell(row=r, column=last_col + 1)
            formula = f"=SUM({get_column_letter(col_offset)}{r}:{get_column_letter(last_col)}{r})"
            total_cell.value = formula
            apply_number_format(total_cell)

        total_row = row_offset + len(sheets_to_process)
        ws_tgt.cell(row=total_row, column=1).value = "合计"
        for col in range(col_offset, last_col + 2):
            formula = f"=SUM({get_column_letter(col)}{row_offset}:{get_column_letter(col)}{total_row - 1})"
            cell = ws_tgt.cell(row=total_row, column=col)
            cell.value = formula
            apply_number_format(cell)

        wb_tgt.save(output_file)
        logging.info(f"输出文件成功: {output_file}")

    except Exception as e:
        logging.exception("执行过程中发生错误")
        raise e
