# main_runner.py：主流程整合脚本
import os
from pathlib import Path
from openpyxl import load_workbook
from modules.collector import collect_summary_values
from src.legacy_runner import run_main_injection 
from inject_modules.table_injector import inject_tables_and_summary
from inject_modules.text_renderer import render_text_template_from_mapping, inject_text_to_excel, inject_summary_values_debug
from modules.mapping_loader import load_mapping_file
from inject_modules.biz import inject_income_expense_all
import logging
import re # <-- 新增：导入 re 模块用于正则表达式匹配

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def run_main():
    project_root = Path(__file__).resolve().parents[1]
    mapping_path = project_root / "data" / "mapping_file.xlsx"
    source_path = project_root / "output" / "output.xlsx"    
    final_path = project_root / "output" / "final_report.xlsx"

    os.makedirs(project_root / "output", exist_ok=True)
    #logging.info("Starting report generation process...")

    # ✅ 构造 alias_dict
    raw_alias_map = load_mapping_file(mapping_path)["subject_alias_map"]
    alias_dict = {}
    for std, aliases in raw_alias_map.items():
        std = std.strip()
        # ✅ 统一转为列表（无论 aliases 是 str 还是 list）
        if not isinstance(aliases, list):
            aliases = [aliases]
        for alias in [std] + aliases:
            alias_dict[alias.strip()] = std

    # Step 0：如果 output.xlsx 不存在，则运行 legacy_runner
    force_regenerate_output = True  # 使用“强制刷新”模式控制是否调用 legacy，则无论 output 是否存在，都调用 legacy_runner 覆盖生成 output.xlsx；
    #正式使用时可设为 False，测试或更新数据时设为 True
    if force_regenerate_output or not source_path.exists():
        #logging.info("Output.xlsx 不存在或强制刷新，启动 legacy_runner 自动注入流程...")
        run_main_injection()
    else:
        logging.info("Output.xlsx 存在，跳过 legacy_runner。")
    # ✅ 第一步：提取 summary_values
    #logging.info("Step 1: Collecting summary values...")
    summary_values = collect_summary_values(mapping_path, source_path)
    #logging.info(f"Collected Summary Values (before date parsing): {summary_values}") # 调整日志，区分前后

    # --- 新增代码：解析审计期间，提取起始日期和终止日期 ---
    audit_period_str = summary_values.get('审计期间')
    start_date = "【未提取】" # 默认值，如果解析失败
    end_date = "【未提取】"   # 默认值，如果解析失败

    if audit_period_str:
        # 尝试匹配 'YYYY年M月-YYYY年M月' 或 'YYYY年M月至YYYY年M月' 格式
        match = re.match(r'(\d{4}年\d{1,2}月)[-至](\d{4}年\d{1,2}月)', audit_period_str)
        if match:
            start_date = match.group(1)
            end_date = match.group(2)
        else:
            # 如果不是标准格式，尝试更简单的分割，例如只取第一个和最后一个日期字符串
            parts = re.split(r'[-至]', audit_period_str)
            if len(parts) == 2:
                start_date = parts[0].strip()
                end_date = parts[1].strip()
            elif len(parts) == 1:
                end_date = parts[0].strip() # 如果只有一个，则视为终止日期
                logging.warning(f"审计期间 '{audit_period_str}' 格式不包含起始和终止日期，只提取到终止日期。")
            else:
                logging.warning(f"无法解析审计期间 '{audit_period_str}' 的格式。")

    summary_values['起始日期'] = start_date
    summary_values['终止日期'] = end_date
    #logging.info(f"Collected Summary Values (after date parsing): {summary_values}") # 调整日志
    # --- 新增代码结束 ---

    # ✅ 补全 summary_values 中的标准字段
    if alias_dict:
        extended = {}
        for k, v in summary_values.items():
            std_key = alias_dict.get(k)
            if std_key and std_key not in summary_values:
                extended[std_key] = v
        summary_values.update(extended)

    # ✅ 第二步：插入三张表和字段 → 保存 final 报表
    #logging.info("Step 2: Injecting tables and summary into final report...")    
    wb_tgt = load_workbook(source_path, data_only=True)
    if not wb_tgt:
        #logging.error(f"无法加载工作簿：{source_path}")
        return
    income_summary_dict = inject_income_expense_all(mapping_path, source_path, wb_tgt)
    try:
        inject_tables_and_summary(            
            output_path=source_path,
            template_path=source_path,
            summary_values=summary_values,
            dest_path=final_path,
            alias_dict=alias_dict,
            mapping_path=mapping_path
        )

        # 再次加载 output.xlsx（防止表格注入后变化）
        wb_tgt = load_workbook(source_path, data_only=True)
        # 收入 / 支出 / 收支结余 提取逻辑
        income_summary_dict = inject_income_expense_all(mapping_path, source_path, wb_tgt)
        # 注入到 summary_values 中
        summary_values.update(income_summary_dict)
        
        #logging.info("Tables and summary injected successfully.")
    except Exception as e:
        logging.error(f"Error during table and summary injection: {e}")
        return

    # ✅ 第三步：渲染文字模板并写入 final
    logging.info("Step 3: Rendering text template and injecting into final report...")
    # 此时 summary_values 已经包含了 '起始日期' 和 '终止日期'
    rendered_text = render_text_template_from_mapping(mapping_path, summary_values, alias_dict)
    #logging.info(f"Rendered Text (first 200 chars): {rendered_text[:200]}...")
    try:
        inject_text_to_excel(final_path, sheet_name="汇总区块", cell="K1", text=rendered_text)
        #logging.info("Text template rendered and injected successfully.")
    except Exception as e:
        logging.error(f"Error during text injection: {e}")

    # ✅ 第四步：调试用途写入 summary_values 到元数据 Sheet
    logging.info("Step 4: Injecting summary values to debug sheet...")
    try:
        inject_summary_values_debug(final_path, summary_values)
        #logging.info("Debug summary values injected successfully.")
    except Exception as e:
        logging.error(f"Error during debug summary values injection: {e}")

    #logging.info(f"✅ 报表已完成写入：{final_path}")
    #logging.info("Report generation process completed.")

if __name__ == '__main__':
    run_main()