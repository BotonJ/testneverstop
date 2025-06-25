# main_logic.py 最终修复版
import os
import logging
from pathlib import Path
from openpyxl import load_workbook
from modules.mapping_loader import load_mapping_file
from modules.fill_yewu import fill_yewu_by_mapping
from modules.fill_balance_anchor import fill_balance_sheet_by_name
from modules.render_header import render_header
from modules.log_utils import log_write
from inject_modules.inject import run_full_injection
from inject_modules.biz import inject_income_expense_all
from inject_modules.balance_utils import get_balance_core_data


def run_main_injection():
    project_root = Path(__file__).resolve().parents[1]
    mapping_path = project_root / "data" / "mapping_file.xlsx"
    mapping = load_mapping_file(mapping_path)
    df_yewu = mapping.get("yewu_line_map")

    wb_src_path = project_root / "data" / "soce.xlsx"
    wb_tgt_path = project_root / "data" / "t.xlsx"
    output_path = project_root / "output" / "output.xlsx"
    log_dir = project_root / "log"
    os.makedirs(log_dir, exist_ok=True)

    logging.basicConfig(filename=log_dir / "auto_log.txt",
                        level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")

    log_balance, log_yewu = [], []
    alias_dict = {a: k for k, v in mapping["subject_alias_map"].items() for a in [k] + v}
    wb_src = load_workbook(wb_src_path, data_only=True)
    wb_tgt = load_workbook(wb_tgt_path,)
    prev_ws_yewu = None

    for sheet_name in wb_src.sheetnames:
        if "资产负债表" in sheet_name:
            year = int(sheet_name[:4])
            ws_src = wb_src[sheet_name]

            ws_balance = wb_tgt.copy_worksheet(wb_tgt["资产负债表"])
            ws_balance.title = f"{year}资产负债表"
            fill_balance_sheet_by_name(ws_src, ws_balance, alias_dict, log_balance, skip_list=[])
            render_header(wb_tgt, sheet_name=ws_balance.title, year=year, header_meta=mapping["header_meta"])

            if f"{year}业务活动表" in wb_src.sheetnames:
                ws_src_yewu = wb_src[f"{year}业务活动表"]
                ws_yewu = wb_tgt.copy_worksheet(wb_tgt["业务活动表"])
                ws_yewu.title = f"{year}业务活动表"
                core_data = get_balance_core_data(ws_balance, mapping["blocks"], mapping["subject_alias_map"])
                net_asset_fallback = {
                    "期初": core_data.get("期初净资产总额", 0),
                    "期末": core_data.get("期末净资产总额", 0)
                }
                fill_yewu_by_mapping(
                    ws_src_yewu,
                    ws_yewu,
                    df_yewu,
                    prev_ws=prev_ws_yewu,
                    net_asset_fallback=net_asset_fallback,
                    log=log_yewu
                )
                render_header(wb_tgt, sheet_name=ws_yewu.title, year=year, header_meta=mapping["header_meta"])
                prev_ws_yewu = ws_yewu

    for tmpl_sheet in ["资产负债表", "业务活动表"]:
        if tmpl_sheet in wb_tgt.sheetnames:
            wb_tgt.remove(wb_tgt[tmpl_sheet])

    wb_tgt.save(output_path)