# table_injector.py：将 table1-3 和 summary 字段插入 output 文件

import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from openpyxl.utils import column_index_from_string, range_boundaries

# def copy_table_from_sheet1(ws_tpl, ws_combined, start_cell, target_col):
#     """
#     从 ws_tpl 中起始单元格开始，向右下自动识别区域，复制值和样式到目标 sheet
#     """
#     start_col_letter = ''.join(filter(str.isalpha, start_cell))
#     start_row = int(''.join(filter(str.isdigit, start_cell)))
#     start_col = column_index_from_string(start_col_letter)

#         # 检测最大行：往下扫描直到遇到连续空白行
#     max_row = start_row
#     empty_row_count = 0
#     while empty_row_count < 2:
#         val = ws_tpl.cell(max_row, start_col).value
#         if val is None:
#             empty_row_count += 1
#         else:
#             empty_row_count = 0
#         max_row += 1
#     max_row -= 2  # 回退至最后一个非空

#     # 检测最大列：往右扫描直到遇到连续空白列
#     max_col = start_col
#     empty_col_count = 0
#     while empty_col_count < 2:
#         val = ws_tpl.cell(start_row, max_col).value
#         if val is None:
#             empty_col_count += 1
#         else:
#             empty_col_count = 0
#         max_col += 1
#     max_col -= 2

#     for i in range(start_row, max_row + 1):
#         for j in range(start_col, max_col + 1):
#             source_cell = ws_tpl.cell(i, j)
#             target_cell = ws_combined.cell(i - start_row + 1, target_col + j - start_col)
#             target_cell.font = source_cell.font.copy()
#             target_cell.border = source_cell.border.copy()
#             target_cell.fill = source_cell.fill.copy()
#             target_cell.number_format = source_cell.number_format
#             target_cell.alignment = source_cell.alignment.copy()
# 强制复制区域（避免 range 失效）
def force_copy(ws_tpl, ws_combined, start_row, start_col, nrows, ncols, target_col):
    for i in range(nrows):
        for j in range(ncols):
            source_cell = ws_tpl.cell(start_row + i, start_col + j)
            target_cell = ws_combined.cell(i + 1, target_col + j)
            target_cell.value = source_cell.value
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = source_cell.alignment.copy()

# 在 inject_tables_and_summary 中使用
force_copy(ws_tpl, ws_combined, start_row=1, start_col=1, nrows=10, ncols=4, target_col=1)   # table1
force_copy(ws_tpl, ws_combined, start_row=1, start_col=6, nrows=10, ncols=4, target_col=9)  # table2
force_copy(ws_tpl, ws_combined, start_row=10, start_col=1, nrows=10, ncols=4, target_col=17)  # table3




def inject_tables_and_summary(output_path, template_path, summary_values, dest_path):
    """
    将模板文件中 Sheet1 的 table 区域 + summary 字段写入 output 文件新建 sheet。
    """
    wb_out = load_workbook(output_path)
    wb_tpl = load_workbook(template_path)

    ws_tpl = wb_tpl["Sheet1"]
    ws_combined = wb_out.create_sheet("汇总区块")

    # ✅ 插入三个表格区域（起始位置 + 列）
    copy_table_from_sheet1(ws_tpl, ws_combined, "A1", 1)    # table1
    copy_table_from_sheet1(ws_tpl, ws_combined, "H1", 9)    # table2
    copy_table_from_sheet1(ws_tpl, ws_combined, "O1", 17)   # table3

    # ✅ 插入 summary_values 字段
    base_row = 30
    base_col = 1
    ws_combined.cell(base_row, base_col, "文字说明字段：")
    for idx, (k, v) in enumerate(summary_values.items()):
        ws_combined.cell(base_row + idx + 1, base_col, k)
        ws_combined.cell(base_row + idx + 1, base_col + 1, v)
    print("🔍 模板 Sheet1 中非空单元格（位置+值）：")
    for row in range(1, 21):
        for col in range(1, 21):
            val = ws_tpl.cell(row, col).value
            if val is not None:
                print(f"{ws_tpl.cell(row, col).coordinate}: {val}")

    wb_out.save(dest_path)
