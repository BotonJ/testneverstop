from openpyxl import load_workbook
from inject_modules.table1 import inject_table1
from inject_modules.table2 import inject_table2
from inject_modules.table3 import inject_table3
from inject_modules.mapping import get_mapping_conf_and_df

def inject_tables_and_summary(output_path, template_path, summary_values, dest_path, alias_dict=None,mapping_path=None):
    wb_src = load_workbook(output_path, data_only=True)
    wb_tgt = load_workbook(template_path)
    ws_tgt = wb_tgt.active

    # ✅ 构造 df_map 配置（来自 mapping_path.xlsx）
    if mapping_path is None:
        raise ValueError("mapping_path 参数缺失")
    mapping_file = mapping_path  # ✅ 使用外部显式传入路径
    conf1, df1 = get_mapping_conf_and_df(mapping_file, "inj1")
    conf2, df2 = get_mapping_conf_and_df(mapping_file, "inj2")
    conf3, df3 = get_mapping_conf_and_df(mapping_file, "inj3")

    # ✅ 注入三张表格
    inject_table1(wb_src, ws_tgt, conf1, df1, log=None)
    inject_table2(wb_src, ws_tgt, conf2, df2, log=None)
    inject_table3(wb_src, ws_tgt, conf3, df3, log=None)

    # ✅ 注入 summary_values 到表格底部（K区）
    base_row = 30
    base_col = 1
    ws_tgt.cell(base_row, base_col, "文字说明字段：")
    for idx, (k, v) in enumerate(summary_values.items()):
        ws_tgt.cell(base_row + idx + 1, base_col, k)
        ws_tgt.cell(base_row + idx + 1, base_col + 1, v if v is not None else "N/A")

    wb_tgt.save(dest_path)
    #print(f"✅ 表格及文字说明写入完成：{dest_path}")
