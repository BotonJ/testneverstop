from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
import logging

summary_fields = {}

def _inject_summary_fuzzy(subjects_whitelist, ws_tgt, sheets_to_process, wb_src, conf):
    row_offset = 2
    col_offset = 2
    subject_col_map = {}
    col_cursor = col_offset

    ws_tgt.cell(row=1, column=1).value = "项目"

    for i, sheet in enumerate(sheets_to_process):
        ws_src = wb_src[sheet]
        d_header = ws_src.cell(row=4, column=4).value
        year_str = str(d_header).strip() if sheet == conf["end_sheet"] else str(d_header).replace("累计数", "").strip()
        row_tgt = row_offset + i
        ws_tgt.cell(row=row_tgt, column=1).value = year_str

        for r in range(5, ws_src.max_row + 1):
            raw_subject = str(ws_src.cell(row=r, column=1).value).strip()
            value = ws_src.cell(row=r, column=4).value

            matched_name = next((std for std in subjects_whitelist if std in raw_subject), None)
            if matched_name and value not in [None, "", 0, 0.0]:
                if matched_name not in subject_col_map:
                    subject_col_map[matched_name] = col_cursor
                    ws_tgt.cell(row=1, column=col_cursor).value = matched_name
                    col_cursor += 1

                col = subject_col_map[matched_name]
                #ws_tgt.cell(row=row_tgt, column=col).value = float(value)以下为测试代码：
                try:
                    cell = ws_tgt.cell(row=row_tgt, column=col)
                    if cell.__class__.__name__ == "MergedCell":
                        raise ValueError(f"单元格 ({row_tgt}, {col}) 是合并单元格，无法赋值")

                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                except Exception as e:
                    logging.error(f"❌ 写入失败：{row_tgt=}, {col=}, value={value}，错误: {e}")
                ws_tgt.cell(row=row_tgt, column=col).number_format = '#,##0.00'

    last_col = max(subject_col_map.values(), default=col_offset)

    for r in range(row_offset, row_offset + len(sheets_to_process)):
        formula = f"=SUM({get_column_letter(col_offset)}{r}:{get_column_letter(last_col)}{r})"
        ws_tgt.cell(row=r, column=last_col + 1).value = formula
        ws_tgt.cell(row=r, column=last_col + 1).number_format = '#,##0.00'
    ws_tgt.cell(row=1, column=last_col + 1).value = "合计"

    total_row = row_offset + len(sheets_to_process)
    ws_tgt.cell(row=total_row, column=1).value = "合计"
    for col in range(col_offset, last_col + 2):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{row_offset}:{col_letter}{total_row - 1})"
        ws_tgt.cell(row=total_row, column=col).value = formula
        ws_tgt.cell(row=total_row, column=col).number_format = '#,##0.00'

def inject_income_expense_all(mapping_file, source_file, wb_tgt):
    raw_df = pd.read_excel(mapping_file, sheet_name="业务活动表汇总注入配置", header=None)
    conf = {}
    data_start = 0
    for i, row in raw_df.iterrows():
        key = str(row[0]).strip() if pd.notna(row[0]) else ""
        val = str(row[1]).strip() if pd.notna(row[1]) else ""
        if key in ["start_sheet", "end_sheet"]:
            conf[key] = val
        if any(str(cell).strip() in ["类型", "科目名称"] for cell in row if pd.notna(cell)):
            data_start = i
            break

    df_mapping = pd.read_excel(mapping_file, sheet_name="业务活动表汇总注入配置", header=data_start)
    income_subjects = df_mapping[df_mapping["类型"] == "收入"]["科目名称"].astype(str).tolist()
    expense_subjects = df_mapping[df_mapping["类型"] == "支出"]["科目名称"].astype(str).tolist()

    wb_src = load_workbook(source_file, data_only=True)
    sheets = wb_src.sheetnames
    s_idx = sheets.index(conf["start_sheet"])
    e_idx = sheets.index(conf["end_sheet"])
    sheets_to_process = [s for s in sheets[s_idx:e_idx+1] if "资产负债表" not in s]

    ws_income = wb_tgt.create_sheet("收入汇总")
    _inject_summary_fuzzy(income_subjects, ws_income, sheets_to_process, wb_src, conf)

    ws_expense = wb_tgt.create_sheet("支出汇总")
    _inject_summary_fuzzy(expense_subjects, ws_expense, sheets_to_process, wb_src, conf)

    summary_fields = extract_income_expense_summary(ws_income, ws_expense)
    return summary_fields

def extract_income_expense_summary(ws_income, ws_expense):
    result = {}

    def get_total(ws):
        for row in ws.iter_rows(min_row=2):
            if row[0].value and "合计" in str(row[0].value):
                return row[-1].value if isinstance(row[-1].value, (int, float)) else 0.0
        return 0.0

    income_total = get_total(ws_income)
    expense_total = get_total(ws_expense)
    balance = round(income_total - expense_total, 2)

    result["收入汇总"] = income_total
    result["支出汇总"] = expense_total
    result["收支结余汇总"] = balance

    return result
