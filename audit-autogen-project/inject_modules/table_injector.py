# File: inject_modules/table_injector.py

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from inject_modules.table1 import inject_table1
from inject_modules.table2 import inject_table2
from inject_modules.table3 import inject_table3
from inject_modules.mapping import get_mapping_conf_and_df
import logging

# Configure logging for debugging in this module
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


# NEW HELPER FUNCTION - 获取实际写入目标单元格的辅助函数
def get_actual_target_cell_for_write(ws, row, col):
    """
    给定工作表和一个单元格的行和列，
    如果该单元格地址落在合并区域内，则返回合并区域的左上角单元格的 (row, col) 元组，
    否则返回原始 (row, col) 元组。
    这确保了我们写入的值始终在合并区域中实际保存值的地方。
    """
    # Convert row, col to A1-style address for logging clarity
    cell_address_a1 = f"{get_column_letter(col)}{row}"

    for merged_range in ws.merged_cells.ranges:
        # Check if the target cell (row, col) is within this merged range
        if row >= merged_range.min_row and row <= merged_range.max_row and \
           col >= merged_range.min_col and col <= merged_range.max_col:

            # Get the top-left coordinates of the merged range
            top_left_row = merged_range.min_row
            top_left_col = merged_range.min_col
            top_left_address_a1 = f"{get_column_letter(top_left_col)}{top_left_row}"

            logging.info(f"单元格 {cell_address_a1} 属于合并区域 {merged_range.coord}. 将写入到左上角单元格: {top_left_address_a1}")
            return top_left_row, top_left_col
    # If not part of any merged region, return the original row, col
    return row, col


def inject_tables_and_summary(output_path, template_path, summary_values, dest_path, alias_dict=None, mapping_path=None):
    logging.info(f"开始注入表格和汇总数据. 源文件: {output_path}, 目标文件: {dest_path}, 模板文件: {template_path}")
    wb_src = load_workbook(output_path, data_only=True)
    wb_tgt = load_workbook(template_path)
    ws_tgt = wb_tgt.active

    # ✅ 构造 df_map 配置（来自 mapping_path.xlsx）
    if mapping_path is None:
        raise ValueError("mapping_path 参数缺失")
    mapping_file = mapping_path  # ✅ 使用外部显式传入路径
    conf1, df1 = get_mapping_conf_and_df(mapping_file, "inj1")
    conf2, df2 = get_mapping_conf_and_df(mapping_file, "inj2")
    conf3, df3 = get_mapping_conf_and_df(mapping_file, "inj3")

    # ✅ 注入三张表格
    logging.info("开始注入 Table 1...")
    inject_table1(wb_src, ws_tgt, conf1, df1, log=None)
    logging.info("Table 1 注入完成.")

    logging.info("开始注入 Table 2...")
    inject_table2(wb_src, ws_tgt, conf2, df2, log=None)
    logging.info("Table 2 注入完成.")

    logging.info("开始注入 Table 3...")
    inject_table3(wb_src, ws_tgt, conf3, df3, log=None)
    logging.info("Table 3 注入完成.")

    # ✅ 注入 summary_values 到表格底部（K区）
    logging.info("开始注入汇总数据 (summary_values)...")
    base_row = 30
    base_col = 1 # Column A

    # Inject header "文字说明字段："
    # Check if this cell (A30) is merged
    actual_header_row, actual_header_col = get_actual_target_cell_for_write(ws_tgt, base_row, base_col)
    try:
        ws_tgt.cell(actual_header_row, actual_header_col, "文字说明字段：")
        logging.info(f"成功注入 '文字说明字段：' 到单元格 {get_column_letter(actual_header_col)}{actual_header_row}")
    except Exception as e:
        logging.error(f"注入 '文字说明字段：' 到单元格 {get_column_letter(actual_header_col)}{actual_header_row} 时出错: {e}")

    for idx, (k, v) in enumerate(summary_values.items()):
        # Calculate target row for current item
        target_row = base_row + idx + 1
        
        # Inject key (column A)
        actual_key_row, actual_key_col = get_actual_target_cell_for_write(ws_tgt, target_row, base_col)
        try:
            ws_tgt.cell(actual_key_row, actual_key_col, k)
            logging.info(f"成功注入 key '{k}' 到单元格 {get_column_letter(actual_key_col)}{actual_key_row}")
        except Exception as e:
            logging.error(f"注入 key '{k}' 到单元格 {get_column_letter(actual_key_col)}{actual_key_row} 时出错: {e}")

        # Inject value (column B)
        actual_value_row, actual_value_col = get_actual_target_cell_for_write(ws_tgt, target_row, base_col + 1)
        try:
            ws_tgt.cell(actual_value_row, actual_value_col, v if v is not None else "N/A")
            logging.info(f"成功注入 value '{v if v is not None else 'N/A'}' for key '{k}' 到单元格 {get_column_letter(actual_value_col)}{actual_value_row}")
        except Exception as e:
            logging.error(f"注入 value '{v if v is not None else 'N/A'}' for key '{k}' 到单元格 {get_column_letter(actual_value_col)}{actual_value_row} 时出错: {e}")

    try:
        wb_tgt.save(dest_path)
        logging.info(f"报告已成功保存到: {dest_path}")
    except Exception as e:
        logging.error(f"保存报告到 {dest_path} 时出错: {e}")