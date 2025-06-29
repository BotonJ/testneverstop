# text_renderer.py：渲染说明文字并注入表格

from jinja2 import Template, Undefined, Environment, FileSystemLoader
from openpyxl import load_workbook
import pandas as pd
import logging # Added for logging

# Configure logging for debugging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

class StrictUndefined(Undefined):
    """
    Custom Undefined class to raise errors for missing variables in Jinja2 templates.
    """
    def __getattr__(self, name):
        raise ValueError(f"Template variable '{self._undefined_name}.{name}' is not defined.")

    def __getitem__(self, key):
        raise ValueError(f"Template variable '{self._undefined_name}[{key}]' is not defined.")

    def __str__(self):
        raise ValueError(f"Template variable '{self._undefined_name}' is not defined.")

def render_text_template_from_mapping(mapping_path, summary_values,alias_dict=None):
    """
    根据 text_mapping 表中标记为“文字模板”的字段，提取模板并渲染。
    """
    df = pd.read_excel(mapping_path, sheet_name="text_mapping")
    df_template = df[df["字段名"] == "文字模板"]
    if df_template.empty:
        logging.warning("No '文字模板' found in text_mapping sheet.")
        return ""
    cell_value = df_template.iloc[0]["模板"]
    if not isinstance(cell_value, str):
        logging.warning(f"Template value is not a string: {type(cell_value)}")
        return ""

    # Prepare summary_values to handle potential None/empty values gracefully in the template
    # Replace None with "【未提取】" or similar placeholder before rendering
    cleaned_summary_values = {
        k: v if v is not None and str(v).strip() != "" else "【未提取】"
        for k, v in summary_values.items()
    }
    if alias_dict:
        alias_extended = {}
        for k, v in cleaned_summary_values.items():
            alias_k = alias_dict.get(k, None)
            if alias_k and alias_k not in cleaned_summary_values:
                alias_extended[alias_k] = v
        cleaned_summary_values.update(alias_extended)
    try:
        # Use a more robust Jinja2 environment
        env = Environment(undefined=StrictUndefined) # Use StrictUndefined for better debugging
        tmpl = env.from_string(cell_value)
        rendered_text = tmpl.render(**cleaned_summary_values)
        return rendered_text
    except Exception as e:
        logging.error(f"【模板渲染失败: {e}】 with summary_values: {cleaned_summary_values}")
        return f"【模板渲染失败: {e}】"


def inject_text_to_excel(file_path, sheet_name="汇总区块", cell="K1", text=""):
    """
    Injects rendered text into a specified cell in an Excel file.
    """
    try:
        wb = load_workbook(file_path)
        # Check if sheet exists, otherwise create it.
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            logging.info(f"Created new sheet: {sheet_name}")

        ws[cell] = text
        wb.save(file_path)
        #logging.info(f"Text successfully injected to {sheet_name}!{cell} in {file_path}")
    except Exception as e:
        logging.error(f"Error injecting text to excel {file_path} - {sheet_name}!{cell}: {e}")


def inject_summary_values_debug(file_path, summary_values, sheet_name="元数据"):
    """
    Injects summary_values into a debug sheet for verification.
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Clear existing data in debug sheet to avoid appending duplicates on re-run
            for row in ws['A1:B' + str(ws.max_row)]:
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(sheet_name)
            #logging.info(f"Created new debug sheet: {sheet_name}")

        ws.cell(1, 1, "字段名")
        ws.cell(1, 2, "对应值")
        for idx, (k, v) in enumerate(summary_values.items()):
            ws.cell(idx + 2, 1, k)
            ws.cell(idx + 2, 2, v)

        wb.save(file_path)
        #logging.info(f"Summary values injected to debug sheet '{sheet_name}' in {file_path}")
    except Exception as e:
        logging.error(f"Error injecting summary values to debug sheet {file_path} - {sheet_name}: {e}")